#!/usr/bin/env python3
"""Run integration tests for all 6 fixes."""
import json
import os
import subprocess
import sys

BASE = "http://localhost:8000"
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
TEST_DATA = os.path.join(PROJECT_ROOT, "test_data")

def curl_json(method, url, data=None, files=None):
    """Run curl and parse JSON response."""
    cmd = ["curl", "-s", "-X", method, url]
    if data and not files:
        cmd.extend(["-H", "Content-Type: application/json", "-d", data])
    if files:
        for f in files:
            cmd.extend(["-F", f"file=@{f}"])
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, cwd=PROJECT_ROOT)
        return json.loads(result.stdout) if result.stdout.strip() else {}
    except Exception as e:
        return {"error": str(e), "raw": result.stdout if 'result' in dir() else ""}

def main():
    print("=" * 60)
    print("INTEGRATION TEST - 6 Fixes")
    print("=" * 60)

    # Create multi_tables.xlsx
    print("\n--- Creating multi_tables.xlsx ---")
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'], ws['B1'] = 'Product', 'Sales'
        ws['A2'], ws['B2'] = 'Widget A', 100
        ws['A3'], ws['B3'] = 'Widget B', 200
        ws['A6'], ws['B6'] = 'Region', 'Revenue'
        ws['A7'], ws['B7'] = 'North', 50000
        ws['A8'], ws['B8'] = 'South', 75000
        os.makedirs(TEST_DATA, exist_ok=True)
        wb.save(os.path.join(TEST_DATA, "multi_tables.xlsx"))
        print("Created multi_tables.xlsx")
    except Exception as e:
        print(f"Could not create multi_tables.xlsx: {e}")

    # Test 1 - Merged cells (skip if no Financial_Projections file)
    print("\n" + "=" * 60)
    print("TEST 1 - Merged cells fix")
    print("=" * 60)
    fin_file = os.path.join(TEST_DATA, "Financial_Projections.xls")
    if not os.path.exists(fin_file):
        fin_file = os.path.join(PROJECT_ROOT, "Financial_Projections.xlsx")
    if not os.path.exists(fin_file):
        print("SKIP: No Financial_Projections file found")
    else:
        r = curl_json("POST", f"{BASE}/workbooks/upload", files=[fin_file])
        if r.get("error"):
            print(f"Upload failed: {r}")
        else:
            print(f"Upload OK: workbook_id={r.get('workbook_id')}")
            r2 = curl_json("GET", f"{BASE}/workbooks/{r.get('workbook_id')}")
            sheets = r2.get("sheets", [])
            print(f"Sheets detected: {len(sheets)}")
            for s in sheets[:2]:
                cols = s.get("columns", [])
                rows = s.get("sample_rows", [])[:2]
                print(f"  Sheet: {s.get('sheet_name')}, cols: {cols[:5]}..., rows sample: {len(rows)}")

    # Test 2 - Formula errors
    print("\n" + "=" * 60)
    print("TEST 2 - Formula error detection")
    print("=" * 60)
    fe = os.path.join(TEST_DATA, "formula_errors.csv")
    r = curl_json("POST", f"{BASE}/workbooks/upload", files=[fe])
    if r.get("error"):
        print(f"Upload failed: {r}")
    else:
        wid = r.get("workbook_id")
        print(f"Upload OK: {wid}")
        rq = curl_json("GET", f"{BASE}/workbooks/{wid}/quality-report")
        issues = rq.get("issues", [])
        formula_issues = [i for i in issues if i.get("rule") == "formula_errors"]
        print(f"Formula error issues: {len(formula_issues)}")
        for fi in formula_issues:
            print(f"  {fi.get('severity')}: {fi.get('message')}")

    # Test 3 - Subtotal detection
    print("\n" + "=" * 60)
    print("TEST 3 - Subtotal row detection")
    print("=" * 60)
    st = os.path.join(TEST_DATA, "subtotals.csv")
    r = curl_json("POST", f"{BASE}/workbooks/upload", files=[st])
    if r.get("error"):
        print(f"Upload failed: {r}")
    else:
        wid = r.get("workbook_id")
        rq = curl_json("GET", f"{BASE}/workbooks/{wid}/quality-report")
        subtotal_issues = [i for i in rq.get("issues", []) if i.get("rule") == "subtotal_rows"]
        print(f"Subtotal issues: {len(subtotal_issues)} - {subtotal_issues}")
        rchat = curl_json("POST", f"{BASE}/chat/query", data=json.dumps({
            "workbook_id": wid,
            "question": "What is the total for q1?"
        }))
        ans = rchat.get("answer", "N/A")
        print(f"Chat answer for 'total for q1': {ans[:200]}...")

    # Test 4 - Multi-row headers
    print("\n" + "=" * 60)
    print("TEST 4 - Multi-row header detection")
    print("=" * 60)
    mh = os.path.join(TEST_DATA, "multi_headers.csv")
    r = curl_json("POST", f"{BASE}/workbooks/upload", files=[mh])
    if r.get("error"):
        print(f"Upload failed: {r}")
    else:
        wid = r.get("workbook_id")
        r2 = curl_json("GET", f"{BASE}/workbooks/{wid}")
        sheets = r2.get("sheets", [])
        for s in sheets:
            cols = s.get("columns", [])
            print(f"Columns: {cols}")

    # Test 5 - Multiple tables
    print("\n" + "=" * 60)
    print("TEST 5 - Multiple tables detection")
    print("=" * 60)
    mt = os.path.join(TEST_DATA, "multi_tables.xlsx")
    if os.path.exists(mt):
        r = curl_json("POST", f"{BASE}/workbooks/upload", files=[mt])
        if r.get("error"):
            print(f"Upload failed: {r}")
        else:
            wid = r.get("workbook_id")
            r2 = curl_json("GET", f"{BASE}/workbooks/{wid}")
            s = r2.get("sheets", [{}])[0]
            print(f"has_multiple_tables: {s.get('has_multiple_tables')}")
            print(f"table_count: {s.get('table_count')}")
            rq = curl_json("GET", f"{BASE}/workbooks/{wid}/quality-report")
            multi = [i for i in rq.get("issues", []) if i.get("rule") == "multiple_tables_on_sheet"]
            print(f"Multiple tables quality issue: {multi}")
    else:
        print("SKIP: multi_tables.xlsx not found")

    # Test 6 - LLM column normalization
    print("\n" + "=" * 60)
    print("TEST 6 - LLM column normalization (workspace)")
    print("=" * 60)
    rw = curl_json("POST", f"{BASE}/workspaces/create", data=json.dumps({"name": "Test CRM/ERP"}))
    wsid = rw.get("workspace_id")
    if not wsid:
        print(f"Workspace create failed: {rw}")
    else:
        print(f"Workspace: {wsid}")
        for fname in ["crm_export.csv", "erp_export.csv"]:
            fp = os.path.join(TEST_DATA, fname)
            ru = curl_json("POST", f"{BASE}/workspaces/{wsid}/upload", files=[fp])
            print(f"  Upload {fname}: {ru.get('status', ru.get('error', 'ok'))}")
        rdet = curl_json("POST", f"{BASE}/workspaces/{wsid}/detect-relationships")
        rels = rdet.get("relationships", [])
        print(f"Relationships detected: {len(rels)}")
        for rel in rels[:5]:
            print(f"  {rel.get('column_1')} <-> {rel.get('column_2')}: conf={rel.get('confidence')}")

    print("\n" + "=" * 60)
    print("DONE")
    print("=" * 60)

if __name__ == "__main__":
    main()
