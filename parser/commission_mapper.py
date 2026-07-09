import os
import json
import sqlite3
import re
from datetime import datetime

CANONICAL_FIELDS = [
    'po_number',
    'dealer_name',
    'invoice_number',
    'invoice_date',
    'sale_amount',
    'commission_rate',
    'commission_amount',
    'comm_credit',
]

MAPPING_PROMPT = """You are a commission report
column mapping expert for a food service rep group.

I will show you column headers and sample rows
from a manufacturer commission report spreadsheet.

Map each canonical field to the EXACT column name
from the spreadsheet headers.

Canonical fields:
- po_number: purchase order number
- dealer_name: dealer or distributor name
- invoice_number: invoice or order number
- invoice_date: date of invoice
- sale_amount: merchandise or net sale value
- commission_rate: commission percentage
- commission_amount: commission dollar amount

Rules:
- Use the EXACT column header string as it appears
- If a field has no matching column use null
- manufacturer: extract from file header or filename
- period: extract month and year from header or filename
- skip_rows_where: describe rows to skip
  e.g. "rows where first column contains Total or Subtotal"

IMPORTANT: For Blodgett reports the correct commission column is always
the TOTAL column (col 23) which equals Origination + Specification + Destination.
Never map commission_amount to Full Comm or Full Comm less rebate
for Blodgett. The Total column is always the authoritative commission.

commission_amount: the FINAL distributed commission amount after
all splits.

When a report has BOTH:
  - A net commission column (e.g. Full Comm less rebate, Net Comm)
  - A Total column that equals the sum of Origination + Specification
    + Destination

Always use the TOTAL column (Orig+Spec+Dest) as commission_amount,
NOT the net commission. The Total column represents the actual
commission distributed to the rep. The net commission column is an
intermediate calculation before distribution.

If no Total/splits exist: use the NET commission amount after any
rebates or deductions.

Common column names for net commission:
  Full Comm less rebate, Net Commission, Commission Net, Net Comm,
  Total (when it equals splits sum), Comm Net

Common column names for GROSS commission (do NOT use these if a net
column exists): Full Comm, Gross Commission, Commission Gross, Gross Comm

commission_rebate: if the report has a rebate or deduction column
map it here. May be labeled: Rebate, Rebate %, Rebate Amount, Deduction

IMPORTANT: Commission amount is always a dollar value. Sale amount
is also a dollar value and is usually larger than commission.
Commission rate is a percentage between 0 and 100. Do not confuse
these three fields. Look carefully at the sample data values
to identify which column contains which type.

- primary_sheet: identify which sheet contains the individual
  transaction line items. This is usually named Input, Data,
  Detail, Transactions, or similar.
- skip_sheets: list any sheets that are summaries, pivots,
  totals, or duplicate the primary sheet. These should be
  skipped to avoid double counting.

Return ONLY valid JSON:
{
  "po_number": "exact column name or null",
  "dealer_name": "exact column name or null",
  "invoice_number": "exact column name or null",
  "invoice_date": "exact column name or null",
  "sale_amount": "exact column name or null",
  "commission_rate": "exact column name or null",
  "commission_amount": "exact column name or null",
  "commission_origination": "exact column name for origination commission or null",
  "commission_specification": "exact column name for specification commission or null",
  "commission_destination": "exact column name for destination commission or null",
  "commission_rebate": "column name or null",
  "manufacturer": "name or null",
  "period": "Month YYYY or null",
  "skip_rows_where": "description or null",
  "primary_sheet": "name of the sheet with individual line items, not summaries or pivots",
  "skip_sheets": ["list of sheet names to skip because they are summaries, pivots, or duplicates of primary_sheet"]
}"""

# Append commission splits rules to the prompt (used in map_columns_with_ai)
MAPPING_PROMPT += """

Commission splits: Some manufacturers split commission into three parts.
Look for columns with these labels (exact match or abbreviation):

Origination: Orig, Originatin, Origination, O Comm, Orig Comm

Specification: Spec, SPEC, Specification, S Comm, Spec Comm

Destination: Dest, DEST, Destination, D Comm, Dest Comm

IMPORTANT: Single letter columns O, S, D are usually status FLAGS not
commission splits. Only map them if they contain dollar amounts, not
single characters like Y/N or status codes.

If split columns exist map all three. If they do not exist set all
three to null. commission_amount should always be the TOTAL commission
regardless of splits."""


def get_headers_and_samples(file_path, max_rows=8):
    """
    Read headers and sample rows using smart
    header detection for Excel (finds real
    header row, not just row 0).
    """
    from pathlib import Path
    ext = Path(file_path).suffix.lower()

    if ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        return _get_excel_samples_smart(
            file_path, max_rows
        )
    elif ext == '.slk':
        return _get_slk_samples_smart(
            file_path, max_rows
        )
    elif ext in ['.csv', '.tsv']:
        return _get_csv_samples(
            file_path, max_rows
        )
    elif ext == '.txt':
        return _get_txt_samples(
            file_path, max_rows
        )
    return []


def _get_excel_samples_smart(file_path, max_rows=8):
    """
    Read Excel headers using smart header
    detection - finds the real header row
    not just row 0.
    """
    import openpyxl

    wb = openpyxl.load_workbook(
        file_path, data_only=True
    )
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect all rows
        all_rows = []
        for row in ws.iter_rows(
            max_row=30, values_only=True
        ):
            cells = [
                str(v).strip()
                if v is not None else ''
                for v in row
            ]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 2:
                all_rows.append(cells)
            if len(all_rows) >= 25:
                break

        if not all_rows:
            continue

        # Find header row - the row with the
        # most string (non-numeric) values
        # that look like column labels
        best_header_idx = 0
        best_score = 0
        best_col_count = 0

        for i, row in enumerate(all_rows[:15]):
            non_empty = [c for c in row if c]
            if not non_empty:
                continue

            # Score: count string values that look like column names
            string_count = 0
            for v in non_empty:
                try:
                    float(str(v).replace(',', ''))
                    continue
                except ValueError:
                    pass
                if re.match(
                    r'\d{1,2}[/-]\d{1,2}', str(v)
                ):
                    continue
                # Skip long report titles (e.g. "Commission Report Jan 2025...")
                if len(str(v)) > 50:
                    continue
                string_count += 1

            score = string_count / max(len(non_empty), 1)
            col_count = len(non_empty)

            if score > best_score or (
                score == best_score and col_count > best_col_count
            ):
                best_score = score
                best_col_count = col_count
                best_header_idx = i

        # Use best header row and next
        # few rows as samples
        header_row = all_rows[best_header_idx]
        sample_rows = all_rows[
            best_header_idx + 1:
            best_header_idx + max_rows
        ]

        results.append({
            'sheet': sheet_name,
            'rows': [header_row] + sample_rows,
            'header_row_idx': best_header_idx
        })

    return results


def _get_slk_samples_smart(file_path, max_rows=8):
    """
    Headers/samples for SYLK (.slk) using the same
    scoring as Excel samples.
    """
    import pandas as pd
    from parser.commission_extractor import load_slk_dataframe

    df = load_slk_dataframe(file_path)
    if df is None or df.empty:
        return []

    all_rows = []
    for _, row in df.head(30).iterrows():
        cells = []
        for v in row:
            if pd.isna(v):
                cells.append('')
            else:
                s = str(v).strip()
                cells.append('' if s.lower() == 'nan' else s)
        if len([c for c in cells if c]) >= 2:
            all_rows.append(cells)
        if len(all_rows) >= 25:
            break

    if not all_rows:
        return []

    best_header_idx = 0
    best_score = 0
    best_col_count = 0

    for i, row in enumerate(all_rows[:15]):
        non_empty = [c for c in row if c]
        if not non_empty:
            continue

        string_count = 0
        for v in non_empty:
            try:
                float(str(v).replace(',', ''))
                continue
            except ValueError:
                pass
            if re.match(
                r'\d{1,2}[/-]\d{1,2}', str(v)
            ):
                continue
            if len(str(v)) > 50:
                continue
            string_count += 1

        score = string_count / max(len(non_empty), 1)
        col_count = len(non_empty)

        if score > best_score or (
            score == best_score and col_count > best_col_count
        ):
            best_score = score
            best_col_count = col_count
            best_header_idx = i

    header_row = all_rows[best_header_idx]
    sample_rows = all_rows[
        best_header_idx + 1:
        best_header_idx + max_rows
    ]

    return [{
        'sheet': 'Sheet1',
        'rows': [header_row] + sample_rows,
        'header_row_idx': best_header_idx,
    }]


def _get_csv_samples(file_path, max_rows):
    import csv
    rows = []
    with open(file_path, 'r',
              encoding='utf-8',
              errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            non_empty = [v for v in row if v.strip()]
            if len(non_empty) >= 2:
                rows.append(row)
            if len(rows) >= max_rows:
                break
    return [{'sheet': 'Sheet1', 'rows': rows}] if rows else []


def _get_txt_samples(file_path, max_rows):
    rows = []
    with open(file_path, 'r',
              encoding='utf-8',
              errors='replace') as f:
        for line in f:
            line = line.strip()
            if len(line) > 10:
                rows.append([line])
            if len(rows) >= max_rows:
                break
    return [{'sheet': 'Sheet1', 'rows': rows}]


def map_columns_with_ai(file_path, file_name):
    """
    Use Gemini to map column headers to
    canonical field names.
    Called ONCE per new manufacturer.
    """
    from google import genai
    from dotenv import load_dotenv
    load_dotenv()

    api_key = os.getenv('GEMINI_API_KEY')
    client = genai.Client(api_key=api_key)

    sheets_data = get_headers_and_samples(
        file_path, max_rows=15
    )

    evidence = f"File name: {file_name}\n\n"
    for sheet in sheets_data:
        evidence += f"Sheet: {sheet['sheet']}\n"
        for i, row in enumerate(sheet['rows']):
            label = "HEADERS" if i == 0 else f"Row {i}"
            evidence += f"{label}: {' | '.join(row)}\n"
        evidence += "\n"

    print(f"  Sending {len(evidence)} chars to Gemini "
          f"for column mapping...")

    response = client.models.generate_content(
        model='gemini-2.5-flash-lite',
        config={'system_instruction': MAPPING_PROMPT},
        contents=evidence
    )

    text = response.text.strip()
    if text.startswith('```'):
        text = text.split('```')[1]
        if text.startswith('json'):
            text = text[4:]
    text = text.strip().rstrip('`').strip()

    mapping = json.loads(text)
    mfr = mapping.get('manufacturer', '') or ''
    mfr = re.sub(r'\s*\(.*?\)\s*$', '', mfr).strip()
    mapping['manufacturer'] = mfr
    print(f"  Column mapping result: {mapping}")
    return mapping


def verify_template_columns(file_path, template):
    """
    Check if template columns still exist
    in the file. Uses fuzzy matching so
    minor column name differences don't
    trigger remapping.
    Returns (is_valid, missing_columns)
    """
    from pathlib import Path

    ext = Path(file_path).suffix.lower()
    sheets_data = get_headers_and_samples(
        file_path, max_rows=2
    )

    if not sheets_data:
        return True, []

    # When template has primary_sheet, verify only that sheet
    # (avoids mismatch with Detail/Voucher having different cols)
    primary = (template.get('primary_sheet') or '').strip().lower()
    skip_sheets = [
        s.strip().lower()
        for s in (template.get('skip_sheets') or [])
        if s
    ]

    all_headers = set()
    for sheet in sheets_data:
        sn = (sheet.get('sheet') or '').strip().lower()
        if primary and sn != primary:
            if ext != '.slk':
                continue
        if skip_sheets and sn in skip_sheets:
            continue
        if sheet['rows']:
            for cell in sheet['rows'][0]:
                if cell and cell.strip():
                    all_headers.add(
                        cell.strip().lower()
                    )
    if not all_headers and primary:
        all_headers = set()
        for sheet in sheets_data:
            if sheet['rows']:
                for cell in sheet['rows'][0]:
                    if cell and cell.strip():
                        all_headers.add(
                            cell.strip().lower()
                        )

    def header_matches(col_name):
        if not col_name:
            return True
        col_lower = col_name.strip().lower()

        # Exact match
        if col_lower in all_headers:
            return True

        # Substring match - template col
        # is contained in an actual header
        for h in all_headers:
            if col_lower in h or h in col_lower:
                return True

        # Word overlap match
        col_words = set(col_lower.split())
        col_words -= {
            'the', 'a', 'an', 'of',
            'and', 'or', 'in', 'at'
        }
        if not col_words:
            return True

        for h in all_headers:
            h_words = set(h.split())
            overlap = col_words & h_words
            if len(overlap) >= max(
                1, len(col_words) * 0.5
            ):
                return True

        return False

    missing = []
    for field in CANONICAL_FIELDS:
        col_name = template.get(field)
        if not col_name:
            continue
        if (
            field == 'commission_amount'
            and ' - ' in str(col_name)
        ):
            parts = str(col_name).split(' - ', 1)
            for p in parts:
                p = p.strip()
                if p and not header_matches(p):
                    missing.append(p)
            continue
        if not header_matches(col_name):
            missing.append(col_name)

    is_valid = len(missing) == 0
    return is_valid, missing


def save_template(
    manufacturer, mapping, db_path
):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS
        commission_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            manufacturer TEXT UNIQUE NOT NULL,
            mapping TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            used_count INTEGER DEFAULT 0
        )
    """)

    now = datetime.now().isoformat()
    cursor.execute("""
        INSERT INTO commission_templates
        (manufacturer, mapping, created_at, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(manufacturer) DO UPDATE SET
            mapping = excluded.mapping,
            updated_at = excluded.updated_at
    """, (
        manufacturer.lower().strip(),
        json.dumps(mapping),
        now, now
    ))

    conn.commit()
    conn.close()
    print(f"  Template saved for: {manufacturer}")


def load_template(manufacturer, db_path):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT mapping FROM commission_templates
            WHERE manufacturer = ?
        """, (manufacturer.lower().strip(),))

        row = cursor.fetchone()

        if row:
            cursor.execute("""
                UPDATE commission_templates
                SET used_count = used_count + 1
                WHERE manufacturer = ?
            """, (manufacturer.lower().strip(),))
            conn.commit()

        conn.close()
        return json.loads(row[0]) if row else None
    except Exception:
        return None


def get_or_create_template(
    file_path, file_name, manufacturer, db_path,
    original_filename=None,
):
    """
    Load existing template or create new one.
    Also verifies template is still valid
    for this file.
    """
    name_for_period = original_filename or file_name
    existing = load_template(manufacturer, db_path)

    if existing:
        is_valid, missing = verify_template_columns(
            file_path, existing
        )

        if is_valid:
            print(f"  Using saved template for: "
                  f"{manufacturer}")

            # Check if primary_sheet exists in this specific file
            primary = existing.get('primary_sheet')
            if primary:
                from pathlib import Path
                ext = Path(file_path).suffix.lower()
                if ext == '.slk':
                    return existing
                if ext in ['.xlsx', '.xlsm',
                           '.xls', '.xlsb']:
                    import openpyxl
                    try:
                        wb = openpyxl.load_workbook(
                            file_path,
                            data_only=True,
                            read_only=True
                        )
                        actual_sheets = wb.sheetnames
                        actual_lower = [
                            s.lower() for s in actual_sheets
                        ]
                        wb.close()

                        if primary.lower() not in actual_lower:
                            from parser.commission_extractor import (
                                extract_period_from_filename,
                                match_sheet_to_filename_period,
                            )
                            fn_period = extract_period_from_filename(
                                name_for_period
                            )
                            resolved = None
                            if fn_period:
                                resolved = (
                                    match_sheet_to_filename_period(
                                        actual_sheets,
                                        fn_period,
                                    )
                                )
                            if resolved:
                                print(
                                    f"  Primary sheet "
                                    f"'{primary}' not in file; "
                                    f"using '{resolved}' "
                                    f"(matches filename period)"
                                )
                                modified = dict(existing)
                                modified['primary_sheet'] = resolved
                                return modified
                            print(f"  Primary sheet "
                                  f"'{primary}' not in "
                                  f"file, using all sheets")
                            # Return template without
                            # primary_sheet constraint
                            modified = dict(existing)
                            modified['primary_sheet'] = None
                            modified['skip_sheets'] = []
                            return modified
                    except Exception:
                        pass

            return existing
        else:
            print(f"  Template outdated - columns "
                  f"missing: {missing}")
            print(f"  Re-mapping with AI...")
    else:
        print(f"  No template for: {manufacturer}")
        print(f"  Creating with AI...")

    mapping = map_columns_with_ai(
        file_path, name_for_period
    )
    save_template(manufacturer, mapping, db_path)
    return mapping


def extract_manufacturer_from_filename(file_name):
    """
    Extract manufacturer name from filename.
    """
    known = [
        'cambro', 'southbend', 'south bend',
        'blodgett', 'blendtec', 'follett',
        'star', 'dormont', 'americanmetalcraft',
        'american metalcraft', 'vulcan',
        'hobart', 'manitowoc', 'true',
        'hoshizaki', 'beverage air',
    ]

    name_lower = file_name.lower()
    name_clean = re.sub(r'[^a-z]', '', name_lower)

    for mfr in known:
        mfr_clean = re.sub(r'[^a-z]', '', mfr)
        if mfr_clean in name_clean:
            return mfr.title()

    # Use filename without date/numbers
    base = file_name.split('.')[0]
    base = re.sub(r'[\d_\-]', ' ', base)
    base = re.sub(r'\s+', ' ', base).strip()

    # Remove common words
    stop_words = [
        'comm', 'commission', 'paid', 'report',
        'fw', 'fwd', 'january', 'february', 'march',
        'april', 'may', 'june', 'july', 'august',
        'september', 'october', 'november', 'december',
        'jan', 'feb', 'mar', 'apr', 'jun', 'jul',
        'aug', 'sep', 'oct', 'nov', 'dec',
        'gabriel', 'group', 'east', 'west',
        '2024', '2025', '2026'
    ]

    words = [
        w for w in base.lower().split()
        if w not in stop_words
        and len(w) > 2
    ]

    result = ' '.join(words).title()

    # Never return generic non-names
    invalid_names = [
        'unknown', 'file', 'report',
        'commission', 'data', ''
    ]
    if result.lower() in invalid_names or len(result) < 3:
        return 'Unknown Manufacturer'

    return result
