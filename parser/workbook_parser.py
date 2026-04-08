"""Parse Excel workbooks into structured data."""
import hashlib
import math
from pathlib import Path
from typing import Any

import chardet
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".xlsb", ".csv", ".tsv", ".ods"}


def _sanitize_for_json(obj: Any) -> Any:
    """Recursively replace NaN/inf with None so JSON serialization succeeds."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    # Handle float NaN and inf (Python and numpy)
    try:
        if isinstance(obj, (int, bool, str)) or obj is None:
            return obj
        if isinstance(obj, float):
            if obj != obj or abs(obj) == float("inf"):
                return None
            return obj
        if pd.isna(obj):
            return None
        if hasattr(obj, "item"):
            v = obj.item()
            if isinstance(v, float) and (v != v or abs(v) == float("inf")):
                return None
    except (TypeError, ValueError, AttributeError):
        pass
    return obj


def _is_numeric_parser(value: Any) -> bool:
    """Helper to check if a value is numeric."""
    if value is None:
        return False
    try:
        s = str(value).replace(",", "").replace("$", "").replace("%", "").strip()
        if not s:
            return False
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def detect_header_row_smart(sheet_data: list, max_scan_rows: int = 10) -> int:
    """
    Smarter header row detection.
    Skip title rows and find first row that looks like a header
    (mostly strings, followed by row with numeric data).
    """
    if not sheet_data:
        return 0

    rows_to_scan = sheet_data[:max_scan_rows]
    best_header_row = 0
    best_score = 0

    for i, row in enumerate(rows_to_scan):
        if isinstance(row, dict):
            values = list(row.values())
        else:
            values = list(row) if hasattr(row, "__iter__") and not isinstance(row, str) else [row]

        non_empty = [v for v in values if v is not None and str(v).strip()]
        if len(non_empty) < 2:
            continue

        string_count = sum(1 for v in non_empty if not _is_numeric_parser(v))
        next_row_numeric = 0
        if i + 1 < len(rows_to_scan):
            next_row = rows_to_scan[i + 1]
            if isinstance(next_row, dict):
                next_values = list(next_row.values())
            else:
                next_values = list(next_row) if hasattr(next_row, "__iter__") and not isinstance(next_row, str) else [next_row]
            next_non_empty = [v for v in next_values if v is not None and str(v).strip()]
            next_row_numeric = sum(1 for v in next_non_empty if _is_numeric_parser(v))

        score = (string_count * 2) + (next_row_numeric * 3) + len(non_empty)
        if score > best_score:
            best_score = score
            best_header_row = i

    return best_header_row


def detect_and_merge_headers(rows_data: list, max_header_rows: int = 3) -> tuple[list, int, list[str] | None]:
    """
    Detects if multiple rows form a combined header and merges them.
    Returns (data_rows, header_row_count, merged_columns or None).
    """
    if not rows_data or len(rows_data) < 2:
        return rows_data, 0, None

    def is_header_row(row_values: list) -> bool:
        non_empty = [v for v in row_values if v is not None and str(v).strip()]
        if not non_empty:
            return False
        numeric_count = 0
        for v in non_empty:
            try:
                float(str(v).replace(",", "").replace("$", "").replace("%", ""))
                numeric_count += 1
            except (ValueError, TypeError):
                pass
        return (numeric_count / len(non_empty)) < 0.3

    header_row_count = 0
    for i in range(min(max_header_rows, len(rows_data))):
        row = rows_data[i]
        row_vals = list(row.values()) if isinstance(row, dict) else list(row)
        if is_header_row(row_vals):
            header_row_count += 1
        else:
            break

    if header_row_count <= 1:
        return rows_data, header_row_count, None

    num_cols = max(
        len(rows_data[i].values()) if isinstance(rows_data[i], dict) else len(rows_data[i])
        for i in range(header_row_count)
    )

    merged_cols: list[str] = []
    for col_idx in range(num_cols):
        parts = []
        for row_idx in range(header_row_count):
            row = rows_data[row_idx]
            if isinstance(row, dict):
                vals = list(row.values())
            else:
                vals = list(row)
            val = vals[col_idx] if col_idx < len(vals) else None
            val_str = str(val).strip() if val is not None and str(val).strip() else ""
            if val_str:
                parts.append(val_str)
        merged = "_".join(parts) if parts else f"col_{col_idx}"
        merged = merged.replace(" ", "_").replace("/", "_").replace("-", "_")
        merged_cols.append(merged)

    remaining_rows = rows_data[header_row_count:]
    return remaining_rows, header_row_count, merged_cols


def detect_multiple_tables(rows_data, columns):
    """
    Detects if a sheet contains multiple
    separate tables either stacked vertically
    or placed side by side.

    Returns a list of table boundaries or None
    if only one table detected.
    """
    if not rows_data or len(rows_data) < 4:
        return None

    tables = []

    def row_is_empty(row):
        if isinstance(row, dict):
            values = list(row.values())
        else:
            values = row
        non_empty = [v for v in values if v is not None and str(v).strip()]
        return len(non_empty) == 0

    def row_fill_ratio(row):
        if isinstance(row, dict):
            values = list(row.values())
        else:
            values = row
        if not values:
            return 0
        non_empty = [v for v in values if v is not None and str(v).strip()]
        return len(non_empty) / len(values)

    # DETECTION 1: Stacked vertical tables
    empty_row_groups = []
    in_empty_group = False
    group_start = None

    for i, row in enumerate(rows_data):
        if row_is_empty(row):
            if not in_empty_group:
                in_empty_group = True
                group_start = i
        else:
            if in_empty_group:
                group_end = i - 1
                gap_size = group_end - group_start + 1
                if gap_size >= 2:
                    empty_row_groups.append({
                        'start': group_start,
                        'end': group_end,
                        'size': gap_size
                    })
                in_empty_group = False

    if empty_row_groups:
        boundaries = [0]
        for gap in empty_row_groups:
            boundaries.append(gap['start'])
            boundaries.append(gap['end'] + 1)
        boundaries.append(len(rows_data))

        for k in range(0, len(boundaries) - 1, 2):
            start = boundaries[k]
            end = boundaries[k + 1]

            if end - start < 2:
                continue

            table_rows = rows_data[start:end]
            fill_ratios = [row_fill_ratio(r) for r in table_rows]
            avg_fill = sum(fill_ratios) / len(fill_ratios) if fill_ratios else 0

            if avg_fill < 0.2:
                continue

            tables.append({
                'type': 'vertical',
                'start_row': start,
                'end_row': end,
                'rows': table_rows
            })

        if len(tables) > 1:
            return tables

    # DETECTION 2: Side by side tables
    if not rows_data:
        return None

    if isinstance(rows_data[0], dict):
        col_names = list(rows_data[0].keys())
    else:
        col_names = list(range(len(rows_data[0])))

    col_fill = {}
    for col in col_names:
        filled = 0
        for row in rows_data:
            if isinstance(row, dict):
                v = row.get(col)
            else:
                v = row[col] if col < len(row) else None
            if v is not None and str(v).strip():
                filled += 1
        col_fill[col] = filled / len(rows_data) if rows_data else 0

    separator_cols = [col for col, fill in col_fill.items() if fill < 0.15]

    if separator_cols:
        all_cols = list(col_names)
        col_groups = []
        current_group = []

        for col in all_cols:
            if col in separator_cols:
                if len(current_group) >= 2:
                    col_groups.append(current_group)
                current_group = []
            else:
                current_group.append(col)

        if len(current_group) >= 2:
            col_groups.append(current_group)

        if len(col_groups) > 1:
            side_tables = []
            for group in col_groups:
                group_rows = []
                for row in rows_data:
                    if isinstance(row, dict):
                        filtered = {k: v for k, v in row.items() if k in group}
                    else:
                        filtered = {
                            group[i]: row[group[i]]
                            for i in range(len(group))
                            if group[i] < len(row)
                        }

                    non_empty = [v for v in filtered.values() if v is not None and str(v).strip()]
                    if non_empty:
                        group_rows.append(filtered)

                if len(group_rows) >= 2:
                    side_tables.append({
                        'type': 'side_by_side',
                        'columns': group,
                        'rows': group_rows
                    })

            if len(side_tables) > 1:
                return side_tables

    return None




def parse_workbook(filepath: str | Path) -> dict[str, Any]:
    """
    Parse a spreadsheet workbook. Detects format by extension and routes to
    the appropriate parser. Returns a normalized JSON shape for all formats.
    """
    filepath = Path(filepath)
    ext = filepath.suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type. Accepted: .xlsx .xlsm .xls .xlsb .csv .tsv .ods"
        )

    workbook_id = hashlib.sha256(str(filepath.resolve()).encode()).hexdigest()[:16]
    file_name = filepath.name

    if ext in (".xlsx", ".xlsm"):
        sheets, formula_nodes = _parse_openpyxl(filepath)
    elif ext == ".xls":
        sheets, formula_nodes = _parse_xlrd(filepath)
    elif ext == ".xlsb":
        sheets, formula_nodes = _parse_pyxlsb(filepath)
    elif ext == ".csv":
        sheets, formula_nodes = _parse_csv(filepath)
    elif ext == ".tsv":
        sheets, formula_nodes = _parse_tsv(filepath)
    elif ext == ".ods":
        sheets, formula_nodes = _parse_ods(filepath)
    else:
        raise ValueError(f"Unsupported extension: {ext}")

    result = {
        "workbook_id": workbook_id,
        "file_name": file_name,
        "sheets": sheets,
        "formula_nodes": formula_nodes,
    }
    return _sanitize_for_json(result)


def expand_merged_cells(worksheet):
    """
    Fills merged cell ranges so every cell
    in the range has the value of the
    top-left cell.

    openpyxl only populates the top-left
    cell of a merged range. All others
    return None. This function fixes that.
    """
    merged_ranges = list(worksheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        max_row = merged_range.max_row
        max_col = merged_range.max_col

        top_left_value = worksheet.cell(row=min_row, column=min_col).value

        worksheet.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col).value = top_left_value


def _parse_openpyxl(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .xlsx and .xlsm using openpyxl."""
    wb = load_workbook(filepath, read_only=False, data_only=False)

    for sheet_name in wb.sheetnames:
        try:
            expand_merged_cells(wb[sheet_name])
        except Exception:
            pass

    sheets: list[dict[str, Any]] = []
    for sheet in wb.worksheets:
        sheet_data = _parse_openpyxl_sheet(sheet)
        if sheet_data["row_count"] == 0:
            sheet_data = _parse_openpyxl_sheet_fallback(filepath, sheet.title, sheet_data["is_hidden"])
        sheets.append(sheet_data)

    formula_cells: list[tuple[str, str, str]] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, Cell) and cell.data_type == "f" and cell.value:
                    formula_cells.append((sheet.title, cell.coordinate, str(cell.value)))

    wb.close()
    formula_nodes = _extract_formula_values(filepath, formula_cells)

    return sheets, formula_nodes


def _parse_openpyxl_sheet(sheet: Worksheet) -> dict[str, Any]:
    """Extract sheet metadata, header row, columns, and sample rows from openpyxl."""
    sheet_name = sheet.title
    is_hidden = sheet.sheet_state in ("hidden", "veryHidden")

    # Read raw rows as list of lists (first 120 rows)
    max_rows_to_read = min(120, sheet.max_row) if sheet.max_row else 0
    if max_rows_to_read == 0:
        return {
            "sheet_name": sheet_name,
            "is_hidden": is_hidden,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    raw_rows: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=1, max_row=max_rows_to_read, values_only=True):
        raw_rows.append(list(row))

    if not raw_rows:
        return {
            "sheet_name": sheet_name,
            "is_hidden": is_hidden,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    # Smart header detection (skip title rows)
    header_start = detect_header_row_smart(raw_rows, max_scan_rows=min(10, len(raw_rows)))
    rows_for_merge = raw_rows[header_start:]

    if not rows_for_merge:
        return {
            "sheet_name": sheet_name,
            "is_hidden": is_hidden,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    # Multi-row header detection and merging
    data_rows, header_row_count, merged_cols = detect_and_merge_headers(
        rows_for_merge, max_header_rows=3
    )

    if merged_cols:
        columns = _normalize_columns(merged_cols)
        data_list = data_rows
    else:
        columns = _normalize_columns(rows_for_merge[0])
        data_list = rows_for_merge[1:]

    if not columns:
        return {
            "sheet_name": sheet_name,
            "is_hidden": is_hidden,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    # Convert data rows to sample_rows (list of dicts)
    sample_rows: list[dict[str, Any]] = []
    rows_to_sample = min(100, len(data_list))

    for i in range(rows_to_sample):
        if i >= len(data_list):
            break
        row_vals = data_list[i]
        row_data: dict[str, Any] = {}
        for col_idx, col_name in enumerate(columns):
            val = row_vals[col_idx] if col_idx < len(row_vals) else None
            if isinstance(val, (int, float, str, bool)) or val is None:
                row_data[col_name] = val
            else:
                row_data[col_name] = str(val)
        sample_rows.append(row_data)

    header_row_idx = header_start + 1  # 1-based for Excel
    row_count = len(data_list)

    sheet_result = {
        "sheet_name": sheet_name,
        "is_hidden": is_hidden,
        "header_row": header_row_idx,
        "columns": columns,
        "sample_rows": sample_rows,
        "row_count": row_count,
    }

    # Check for multiple tables
    tables = detect_multiple_tables(sample_rows, columns)

    if tables and len(tables) > 1:
        largest_table = max(tables, key=lambda t: len(t.get("rows", [])))
        sheet_result["sample_rows"] = largest_table.get("rows", sample_rows)
        sheet_result["row_count"] = len(sheet_result["sample_rows"])
        if largest_table.get("columns"):
            sheet_result["columns"] = largest_table["columns"]

        additional_tables = [
            {
                "table_index": i,
                "type": t.get("type"),
                "row_count": len(t.get("rows", [])),
                "start_row": t.get("start_row", 0),
            }
            for i, t in enumerate(tables)
            if t != largest_table
        ]
        sheet_result["has_multiple_tables"] = True
        sheet_result["table_count"] = len(tables)
        sheet_result["additional_tables"] = additional_tables
        print(
            f"Sheet '{sheet_name}': detected {len(tables)} tables "
            f"({tables[0].get('type', 'unknown')} layout)"
        )
    else:
        sheet_result["has_multiple_tables"] = False
        sheet_result["table_count"] = 1

    return sheet_result


def _parse_openpyxl_sheet_fallback(
    filepath: Path, sheet_name: str, is_hidden: bool
) -> dict[str, Any]:
    """
    Fallback when primary parse returns 0 rows: open with data_only=True,
    use first row as header, iterate all rows without minimum non-empty filter.
    Returns whatever text content exists (e.g. labels).
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return {
                "sheet_name": sheet_name,
                "is_hidden": is_hidden,
                "header_row": 0,
                "columns": [],
                "sample_rows": [],
                "row_count": 0,
            }

        header_row = [v for v in rows[0]]
        columns = _normalize_columns(header_row)
        row_count = len(rows) - 1

        sample_rows: list[dict[str, Any]] = []
        rows_to_sample = min(100, row_count)
        for i in range(rows_to_sample):
            row_vals = rows[i + 1] if i + 1 < len(rows) else []
            row_data: dict[str, Any] = {}
            for j, col_name in enumerate(columns):
                val = row_vals[j] if j < len(row_vals) else None
                if isinstance(val, (int, float, str, bool)) or val is None:
                    row_data[col_name] = val
                else:
                    row_data[col_name] = str(val)
            sample_rows.append(row_data)

        return {
            "sheet_name": sheet_name,
            "is_hidden": is_hidden,
            "header_row": 1,
            "columns": columns,
            "sample_rows": sample_rows,
            "row_count": row_count,
        }
    finally:
        wb.close()


def _extract_formula_values(
    filepath: Path,
    formula_cells: list[tuple[str, str, str]],
) -> list[dict[str, Any]]:
    """Open workbook once in data_only mode; return formula nodes with computed values."""
    if not formula_cells:
        return []

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        nodes: list[dict[str, Any]] = []
        for sheet_name, coordinate, formula_str in formula_cells:
            ws = wb[sheet_name]
            cell = ws[coordinate]
            nodes.append({
                "cell": coordinate,
                "formula": formula_str,
                "value": cell.value,
            })
        return nodes
    finally:
        wb.close()


def _parse_xlrd(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .xls using xlrd. No formula strings available, only computed values."""
    import xlrd

    wb = xlrd.open_workbook(str(filepath))
    sheets: list[dict[str, Any]] = []

    for sheet in wb.sheets():
        sheet_data = _parse_xlrd_sheet(sheet)
        sheets.append(sheet_data)

    # xlrd does not expose formula strings; formula_nodes empty
    formula_nodes: list[dict[str, Any]] = []
    return sheets, formula_nodes


def _parse_xlrd_sheet(sheet) -> dict[str, Any]:
    """Extract sheet structure from xlrd sheet."""
    sheet_name = sheet.name
    nrows, ncols = sheet.nrows, sheet.ncols

    header_row_idx = 0
    header_row = []
    for r in range(nrows):
        row_vals = [sheet.cell_value(r, c) for c in range(ncols)]
        non_empty = sum(1 for v in row_vals if v is not None and str(v).strip())
        if non_empty >= 3:
            header_row_idx = r + 1  # 1-based
            header_row = row_vals
            break

    if not header_row:
        return {
            "sheet_name": sheet_name,
            "is_hidden": False,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    columns = _normalize_columns(header_row)
    row_count = max(0, nrows - header_row_idx)

    sample_rows: list[dict[str, Any]] = []
    rows_to_sample = min(100, row_count)
    for i in range(rows_to_sample):
        row_idx = header_row_idx + i
        row_data: dict[str, Any] = {}
        for j, col_name in enumerate(columns):
            if j < ncols:
                val = sheet.cell_value(row_idx, j)
                if isinstance(val, (int, float, str, bool)) or val is None:
                    row_data[col_name] = val
                else:
                    row_data[col_name] = str(val)
            else:
                row_data[col_name] = None
        sample_rows.append(row_data)

    return {
        "sheet_name": sheet_name,
        "is_hidden": False,
        "header_row": header_row_idx,
        "columns": columns,
        "sample_rows": sample_rows,
        "row_count": row_count,
    }


def _parse_pyxlsb(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .xlsb using pyxlsb. No formula strings available, only computed values."""
    from pyxlsb import open_workbook

    sheets: list[dict[str, Any]] = []
    formula_nodes: list[dict[str, Any]] = []

    with open_workbook(str(filepath)) as wb:
        for idx, sheet_name in enumerate(wb.sheets, start=1):
            with wb.get_sheet(idx) as sheet:
                sheet_data = _parse_pyxlsb_sheet(sheet_name, sheet)
                sheets.append(sheet_data)

    return sheets, formula_nodes


def _parse_pyxlsb_sheet(sheet_name: str, sheet) -> dict[str, Any]:
    """Extract sheet structure from pyxlsb sheet."""
    rows_list = list(sheet.rows())
    if not rows_list:
        return {
            "sheet_name": sheet_name,
            "is_hidden": False,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    header_row_idx = 0
    header_row: list[Any] = []
    for r_idx, row in enumerate(rows_list):
        vals = [c.v if hasattr(c, "v") else None for c in row]
        non_empty = sum(1 for v in vals if v is not None and str(v).strip())
        if non_empty >= 3:
            header_row_idx = r_idx + 1
            header_row = vals
            break

    if not header_row:
        return {
            "sheet_name": sheet_name,
            "is_hidden": False,
            "header_row": 0,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    columns = _normalize_columns(header_row)
    row_count = max(0, len(rows_list) - header_row_idx)

    sample_rows: list[dict[str, Any]] = []
    rows_to_sample = min(100, row_count)
    max_cols = len(columns)

    for i in range(rows_to_sample):
        row_idx = header_row_idx + i
        if row_idx >= len(rows_list):
            break
        row_cells = rows_list[row_idx]
        row_data: dict[str, Any] = {}
        for j, col_name in enumerate(columns):
            val = (row_cells[j].v if j < len(row_cells) else None) if j < len(row_cells) else None
            if isinstance(val, (int, float, str, bool)) or val is None:
                row_data[col_name] = val
            else:
                row_data[col_name] = str(val)
        sample_rows.append(row_data)

    return {
        "sheet_name": sheet_name,
        "is_hidden": False,
        "header_row": header_row_idx,
        "columns": columns,
        "sample_rows": sample_rows,
        "row_count": row_count,
    }


# Excel formula error values (for CSV na_values and error detection)
CSV_ERROR_VALUES = {
    "#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NULL!", "#NUM!",
    "#NAME?", "#ERROR!",
}


def _parse_csv(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .csv with chardet encoding detection. Single sheet 'Sheet1'.
    CSV files always have exactly ONE header row. No multi-row header merging.
    """
    raw = filepath.read_bytes()
    detected = chardet.detect(raw)
    encoding = detected.get("encoding") or "utf-8"

    # Read CSV - first row is header (pandas default). No detect_and_merge_headers for CSV.
    df = pd.read_csv(
        filepath,
        encoding=encoding,
        na_values=list(CSV_ERROR_VALUES),
        keep_default_na=False,
    )

    if df.empty:
        sheet = {
            "sheet_name": "Sheet1",
            "is_hidden": False,
            "header_row": 1,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
            "formula_error_cells": [],
            "formula_error_count": 0,
        }
        return [sheet], []

    # Simple: pandas already used first row as header
    columns = _normalize_columns(list(df.columns.astype(str)))
    df.columns = columns

    # Detect formula errors in data cells (scan after header)
    formula_error_cells: list[dict[str, Any]] = []
    for row_idx, row in df.iterrows():
        for col_idx, col_name in enumerate(columns):
            if col_idx < len(row):
                val = row.iloc[col_idx]
                if val is not None and str(val).strip() in CSV_ERROR_VALUES:
                    formula_error_cells.append({
                        "column": col_name,
                        "row": int(row_idx),
                        "error": str(val).strip(),
                    })

    # Convert NaN to None for JSON serialization
    sample_rows_raw = df.head(100).to_dict(orient="records")
    sample_rows: list[dict[str, Any]] = []
    for row in sample_rows_raw:
        clean_row: dict[str, Any] = {}
        for k, v in row.items():
            if v is not None:
                try:
                    if pd.isna(v):
                        v = None
                except (TypeError, ValueError):
                    pass
            clean_row[k] = v
        sample_rows.append(clean_row)

    sheet = {
        "sheet_name": "Sheet1",
        "is_hidden": False,
        "header_row": 1,
        "columns": columns,
        "sample_rows": sample_rows,
        "row_count": len(df),
        "formula_error_cells": formula_error_cells,
        "formula_error_count": len(formula_error_cells),
    }
    return [sheet], []


def _parse_tsv(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .tsv. Single sheet 'Sheet1'."""
    df = pd.read_csv(filepath, sep="\t")
    sheet = _dataframe_to_sheet(df, "Sheet1")
    return [sheet], []


def _parse_ods(filepath: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse .ods using pandas with odf engine."""
    xl = pd.ExcelFile(filepath, engine="odf")
    sheets: list[dict[str, Any]] = []
    for name in xl.sheet_names:
        df = xl.parse(name)
        sheet = _dataframe_to_sheet(df, name)
        sheets.append(sheet)
    return sheets, []


def _dataframe_to_sheet(df: pd.DataFrame, sheet_name: str) -> dict[str, Any]:
    """Convert a DataFrame to the normalized sheet format."""
    if df.empty:
        return {
            "sheet_name": sheet_name,
            "is_hidden": False,
            "header_row": 1,
            "columns": [],
            "sample_rows": [],
            "row_count": 0,
        }

    columns = _normalize_columns(list(df.columns.astype(str)))
    df.columns = columns
    row_count = len(df)
    sample_rows = df.head(100).to_dict(orient="records")

    return {
        "sheet_name": sheet_name,
        "is_hidden": False,
        "header_row": 1,
        "columns": columns,
        "sample_rows": sample_rows,
        "row_count": row_count,
    }


def _col_index_to_letter(idx: int) -> str:
    """Convert 0-based column index to Excel column letter(s): 0->A, 1->B, ..., 26->AA."""
    n = idx + 1
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _normalize_columns(header_row: list[Any]) -> list[str]:
    """Use actual cell value when non-empty. Only use Excel letter (A, B, C) when cell is None or empty."""
    columns = []
    seen = set()
    for i, v in enumerate(header_row):
        raw = "" if v is None else v
        val = str(raw).strip()
        if not val:
            val = _col_index_to_letter(i)
        if val in seen:
            val = f"{val}_{i}"
        seen.add(val)
        columns.append(val)
    return columns
