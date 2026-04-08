import os
import json
import re
import hashlib
import sqlite3
from datetime import date, datetime
from pathlib import Path

# ─────────────────────────────────────
# FILE READERS
# ─────────────────────────────────────


def read_pdf(file_path):
    """Extract text from PDF commission report."""
    try:
        import pdfplumber
        text_pages = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text_pages.append(
                        f"--- PAGE {i+1} ---\n{page_text}"
                    )
        return '\n\n'.join(text_pages)
    except Exception as e:
        raise ValueError(f"PDF read error: {e}")


def read_excel(file_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            file_path, data_only=True
        )
        
        all_text = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(
                f"--- SHEET: {sheet_name} ---"
            )
            
            row_count = 0
            prev_row_text = None
            
            for row in ws.iter_rows(
                values_only=True
            ):
                non_empty = [
                    v for v in row
                    if v is not None
                    and str(v).strip()
                ]
                
                if len(non_empty) < 2:
                    continue
                
                row_text = '|'.join([
                    str(v).strip()
                    for v in row
                    if v is not None
                    and str(v).strip()
                ])
                
                # Skip exact duplicates
                if row_text == prev_row_text:
                    continue
                
                prev_row_text = row_text
                all_text.append(row_text)
                row_count += 1
                
                if row_count >= 3000:
                    all_text.append(
                        '[truncated at 3000 rows]'
                    )
                    break
        
        result = '\n'.join(all_text)
        print(f"  Excel extracted: "
              f"{len(result)} chars")
        return result
    
    except Exception as e:
        try:
            import xlrd
            wb = xlrd.open_workbook(str(file_path))
            all_text = []
            for sheet in wb.sheets():
                all_text.append(
                    f"--- SHEET: {sheet.name} ---"
                )
                prev_row = None
                for row_idx in range(
                    min(sheet.nrows, 3000)
                ):
                    row = sheet.row_values(row_idx)
                    non_empty = [
                        str(v).strip()
                        for v in row
                        if str(v).strip()
                    ]
                    if len(non_empty) < 2:
                        continue
                    row_text = '|'.join(non_empty)
                    if row_text == prev_row:
                        continue
                    prev_row = row_text
                    all_text.append(row_text)
            return '\n'.join(all_text)
        except Exception as e2:
            raise ValueError(
                f"Excel read error: {e} / {e2}"
            )


def read_txt(file_path):
    """Read text/CSV commission report."""
    try:
        with open(file_path, 'r', 
                  encoding='utf-8', 
                  errors='replace') as f:
            return f.read()
    except Exception as e:
        raise ValueError(f"TXT read error: {e}")


def _slk_manual_parse_to_dataframe(file_path):
    """
    Parse SYLK (SLK) text into a pandas DataFrame.
    Cell records look like: C;X1;Y2;K"value" (X/Y order may vary).
    """
    import pandas as pd

    cells = {}
    max_row = 0
    max_col = 0

    with open(file_path, 'r', encoding='latin-1', errors='replace') as f:
        for line in f:
            line = line.strip()
            if not line.startswith('C;'):
                continue
            parts = [p.strip() for p in line.split(';')]
            x = y = None
            val = ''
            for p in parts[1:]:
                if not p:
                    continue
                tag = p[0].upper()
                rest = p[1:]
                if tag == 'X':
                    digits = ''.join(c for c in rest if c.isdigit())
                    if digits:
                        try:
                            x = int(digits)
                        except ValueError:
                            pass
                elif tag == 'Y':
                    digits = ''.join(c for c in rest if c.isdigit())
                    if digits:
                        try:
                            y = int(digits)
                        except ValueError:
                            pass
                elif tag == 'K':
                    val = rest
                    if len(val) >= 2 and val[0] == '"' and val[-1] == '"':
                        val = val[1:-1]
            if x is not None and y is not None and x > 0 and y > 0:
                cells[(y, x)] = val
                max_row = max(max_row, y)
                max_col = max(max_col, x)

    if max_row < 1 or max_col < 1:
        return None

    grid = []
    for r in range(1, max_row + 1):
        row = [cells.get((r, c), '') for c in range(1, max_col + 1)]
        grid.append(row)

    if not grid:
        return None

    hdr = [
        str(c).strip() if str(c).strip() else f'Column{i}'
        for i, c in enumerate(grid[0])
    ]
    if len(grid) == 1:
        return pd.DataFrame(columns=hdr)
    return pd.DataFrame(grid[1:], columns=hdr)


def load_slk_dataframe(file_path):
    """
    Load an SLK file as a single-sheet DataFrame.
    Tries xlrd/pandas first, then manual SYLK parse.
    """
    import pandas as pd

    try:
        df = pd.read_excel(file_path, engine='xlrd')
        if df is not None and not df.empty:
            return df
    except Exception:
        pass

    return _slk_manual_parse_to_dataframe(file_path)


def read_slk(file_path):
    """
    Read SLK (Symbolic Link / SYLK) into pipe-separated text
    similar to read_excel output (for LLM / hash).
    """
    import pandas as pd

    try:
        df = pd.read_excel(file_path, engine='xlrd')
    except Exception:
        df = _slk_manual_parse_to_dataframe(file_path)

    if df is None or df.empty:
        return ''

    lines = ['--- SHEET: Sheet1 ---']
    prev_row_text = None
    row_count = 0
    for _, row in df.iterrows():
        vals = []
        for v in row:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = str(v).strip()
            if s and s.lower() != 'nan':
                vals.append(s)
        if len(vals) < 2:
            continue
        row_text = '|'.join(vals)
        if row_text == prev_row_text:
            continue
        prev_row_text = row_text
        lines.append(row_text)
        row_count += 1
        if row_count >= 3000:
            lines.append('[truncated at 3000 rows]')
            break

    result = '\n'.join(lines)
    print(f"  SLK extracted: {len(result)} chars")
    return result


def read_commission_file(file_path):
    """
    Universal file reader.
    Returns raw text content regardless of format.
    Also returns document hash for dedup.
    """
    path = Path(file_path)
    ext = path.suffix.lower()
    
    if ext == '.pdf':
        content = read_pdf(file_path)
    elif ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        content = read_excel(file_path)
    elif ext == '.slk':
        content = read_slk(file_path)
    elif ext in ['.csv', '.txt', '.tsv']:
        content = read_txt(file_path)
    else:
        raise ValueError(
            f"Unsupported file format: {ext}"
        )
    
    doc_hash = hashlib.sha256(
        content.encode('utf-8', errors='replace')
    ).hexdigest()
    
    return {
        'content': content,
        'hash': doc_hash,
        'file_name': path.name,
        'file_type': ext,
        'char_count': len(content)
    }


# ─────────────────────────────────────
# GEMINI EXTRACTION
# ─────────────────────────────────────

EXTRACTION_PROMPT = """You are a commission report 
extraction expert for a food service rep group.

Extract ALL commission line items from this report.
The report may be from any manufacturer in any 
format — tabular rows, multi-line entries, 
PDF text, Excel data, or fixed-width text.

For EACH line item extract these fields:
  po_number: the purchase order number.
    In Follett/Gabriel PDFs this is labeled PO # and appears on the
    second line of each transaction block (e.g. SP00007011, ODS0025074,
    245030FH0009). It is NOT the Order # which is a different field
    on the first line.
  dealer_name: name of dealer or distributor
  invoice_number: the INVOICE number, not the order number.
    In some reports there are two numbers:
      Order # (the internal order reference)
      Invoice # (the billing invoice number, often shown in blue or
                 as a separate field)
    Always prefer the Invoice # over Order #.
    In Follett/Gabriel Group PDFs the layout is:
      Date Invoiced   Order #
      PO #            Invoice #    Bill Name
      PO Date
      Orig Ship       Orig Inv#    Dealer
    The Invoice # is the second number shown (e.g. 1208638 not 1224835).
    If only one number exists use that.
  invoice_date: date as YYYY-MM-DD
  sale_amount: merchandise/sale value as number
  commission_rate: commission % as decimal (5% = 5.0)
  commission_amount: commission dollars as number
  manufacturer: manufacturer name from report header

IMPORTANT RULES:
- Extract individual line items NOT summary rows
- Skip rows labeled Total, Subtotal, Grand Total
- Skip rows where commission_amount is 0 AND 
  sale_amount is 0
- Multiple line items CAN share the same po_number
- If po_number is missing or blank set to null
- Dates MUST be YYYY-MM-DD format
- Amounts MUST be numbers not strings
- Negative amounts are credits — keep them negative
- commission_rate should be a number like 5.0 not 0.05

Return ONLY this JSON structure, no markdown:
{
  "manufacturer": "name from report header",
  "period": "extract exact month and year from the report header or filename. Format as Month YYYY. Example: February 2025. Do not guess - use what is explicitly stated in the document.",
  "currency": "USD",
  "line_items": [
    {
      "po_number": "string or null",
      "dealer_name": "string",
      "invoice_number": "string",
      "invoice_date": "YYYY-MM-DD or null",
      "sale_amount": 0.00,
      "commission_rate": 0.00,
      "commission_amount": 0.00,
      "manufacturer": "string"
    }
  ]
}"""


def chunk_content(content, max_chars=25000):
    """
    Split large documents into chunks.
    Uses 25000 chars (up from 10000) to
    reduce number of API calls.
    """
    if len(content) <= max_chars:
        return [content]

    chunks = []

    if '--- PAGE ' in content:
        # PDF: split by pages
        pages = content.split('--- PAGE ')
        current_chunk = ''
        for page in pages:
            if len(current_chunk) + len(page) > max_chars:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = '--- PAGE ' + page
            else:
                current_chunk += (
                    '--- PAGE '
                    if current_chunk else ''
                ) + page
        if current_chunk:
            chunks.append(current_chunk)

    elif '--- SHEET:' in content:
        # Excel: split by sheets first
        # then by lines if sheet is too large
        sheets = content.split('--- SHEET:')
        current_chunk = ''

        for sheet in sheets:
            sheet_content = (
                '--- SHEET:' + sheet
                if sheet and not sheet.startswith('---')
                else sheet
            )

            if len(sheet_content) > max_chars:
                # Large sheet - split by lines
                if current_chunk:
                    chunks.append(current_chunk)
                    current_chunk = ''

                lines = sheet_content.split('\n')
                line_chunk = []
                line_size = 0

                for line in lines:
                    line_len = len(line) + 1
                    if line_size + line_len > max_chars:
                        if line_chunk:
                            chunks.append(
                                '\n'.join(line_chunk)
                            )
                        line_chunk = [line]
                        line_size = line_len
                    else:
                        line_chunk.append(line)
                        line_size += line_len

                if line_chunk:
                    chunks.append(
                        '\n'.join(line_chunk)
                    )

            elif len(current_chunk) + len(sheet_content) > max_chars:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = sheet_content

            else:
                current_chunk += sheet_content

        if current_chunk:
            chunks.append(current_chunk)

    else:
        # Text: split by lines
        lines = content.split('\n')
        current_chunk = []
        current_size = 0

        for line in lines:
            line_size = len(line) + 1
            if current_size + line_size > max_chars:
                if current_chunk:
                    chunks.append(
                        '\n'.join(current_chunk)
                    )
                current_chunk = [line]
                current_size = line_size
            else:
                current_chunk.append(line)
                current_size += line_size

        if current_chunk:
            chunks.append('\n'.join(current_chunk))

    return chunks


def extract_with_gemini(content, file_name):
    """
    Use Gemini to extract commission records
    from raw document content.
    Handles large documents by chunking.
    """
    from google import genai
    
    api_key = os.getenv('GEMINI_API_KEY')
    if not api_key:
        raise ValueError("GEMINI_API_KEY not set")
    
    client = genai.Client(api_key=api_key)
    
    chunks = chunk_content(content)
    all_line_items = []
    manufacturer = None
    period = None
    
    for i, chunk in enumerate(chunks):
        print(f"  Processing chunk {i+1}/{len(chunks)} "
              f"({len(chunk)} chars)...")
        
        user_message = (
            f"File name: {file_name}\n"
            f"This file name may contain the period "
            f"(month/year) of the report.\n\n"
            f"Extract commission data "
            f"(chunk {i+1} of {len(chunks)}):\n\n"
            f"{chunk}"
        )
        
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash-lite',
                config={
                    'system_instruction': EXTRACTION_PROMPT
                },
                contents=user_message
            )
            
            text = response.text.strip()
            if text.startswith('```'):
                text = text.split('```')[1]
                if text.startswith('json'):
                    text = text[4:]
            text = text.strip().rstrip('`').strip()
            
            result = json.loads(text)
            
            if not manufacturer:
                manufacturer = result.get('manufacturer')
            if not period:
                period = result.get('period')
            
            items = result.get('line_items', [])
            all_line_items.extend(items)
            print(f"  Extracted {len(items)} items "
                  f"from chunk {i+1}")
        
        except Exception as e:
            print(f"  Chunk {i+1} error: {e}")
            continue
    
    return {
        'manufacturer': manufacturer,
        'period': period,
        'currency': 'USD',
        'line_items': all_line_items,
        'total_extracted': len(all_line_items)
    }


# ─────────────────────────────────────
# VALIDATION AND CONFIDENCE SCORING
# ─────────────────────────────────────

def clean_amount(val):
    """Convert any amount value to float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    # Remove currency symbols and commas
    cleaned = str(val).strip()
    cleaned = cleaned.replace('$', '')
    cleaned = cleaned.replace(',', '')
    cleaned = cleaned.replace('(', '-')
    cleaned = cleaned.replace(')', '')
    cleaned = cleaned.strip()
    if not cleaned or cleaned == '-':
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def normalize_date(date_val):
    """
    Try multiple date formats and return
    YYYY-MM-DD or None.
    """
    if date_val is None:
        return None

    try:
        import pandas as pd
        if pd.isna(date_val):
            return None
    except (TypeError, ImportError, ValueError):
        pass

    if isinstance(date_val, date):
        return date_val.strftime('%Y-%m-%d')

    # Excel serial number (avoid small ints that look like years)
    if isinstance(date_val, (int, float)) and not isinstance(
        date_val, bool
    ):
        if 1 < float(date_val) < 1_000_000:
            try:
                from openpyxl.utils.datetime import from_excel
                dt = from_excel(date_val)
                if isinstance(dt, datetime):
                    return dt.strftime('%Y-%m-%d')
                if isinstance(dt, date):
                    return dt.strftime('%Y-%m-%d')
            except (TypeError, ValueError, ImportError):
                pass

    date_str = str(date_val).strip()
    if not date_str:
        return None

    # Plain YYYY-MM-DD or ISO with time / timezone
    iso_date = re.match(
        r'^(\d{4}-\d{2}-\d{2})',
        date_str,
    )
    if iso_date:
        return iso_date.group(1)

    formats = [
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%m/%d/%y',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%m-%d-%Y',
        '%d-%m-%Y',
        '%B %d, %Y',
        '%b %d, %Y',
        '%d %B %Y',
        '%m/%d/%Y %H:%M',
        '%m/%d/%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%S.%f',
        '%d-%b-%Y',
        '%Y%m%d',
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue

    try:
        from dateutil import parser as dateutil_parser
        dt = dateutil_parser.parse(date_str)
        return dt.strftime('%Y-%m-%d')
    except (ImportError, ValueError, TypeError, OverflowError):
        pass

    return None


def extract_period_from_filename(file_name):
    """Extract period (Month YYYY) from filename."""
    months = {
        'january': 'January', 'jan': 'January',
        'february': 'February', 'feb': 'February',
        'march': 'March', 'mar': 'March',
        'april': 'April', 'apr': 'April',
        'may': 'May',
        'june': 'June', 'jun': 'June',
        'july': 'July', 'jul': 'July',
        'august': 'August', 'aug': 'August',
        'september': 'September', 'sep': 'September',
        'october': 'October', 'oct': 'October',
        'november': 'November', 'nov': 'November',
        'december': 'December', 'dec': 'December',
    }

    name_lower = file_name.lower()

    # Find month
    found_month = None
    for key, val in months.items():
        if key in name_lower:
            found_month = val
            break

    # Find year
    year_match = re.search(r'20\d{2}', file_name)
    found_year = year_match.group() if year_match else None

    # Handle "Jul 25" style (25 = 2025)
    if not found_year:
        year_match = re.search(
            r'[\s_\-](\d{2})[\s_\-]',
            file_name
        )
        if year_match:
            yr = int(year_match.group(1))
            if 20 <= yr <= 30:
                found_year = f"20{yr:02d}"

    # Handle MMYY or YYMM format like "0126"
    base_name = os.path.basename(file_name)
    if not found_year:
        mmyy = re.search(
            r'(?<![0-9])(\d{2})(\d{2})(?![0-9])',
            base_name
        )
        if mmyy:
            p1 = int(mmyy.group(1))
            p2 = int(mmyy.group(2))
            if 1 <= p1 <= 12 and 20 <= p2 <= 30:
                found_year = f"20{p2:02d}"
                if not found_month:
                    months_short = {
                        '01': 'January', '02': 'February',
                        '03': 'March', '04': 'April',
                        '05': 'May', '06': 'June',
                        '07': 'July', '08': 'August',
                        '09': 'September', '10': 'October',
                        '11': 'November', '12': 'December'
                    }
                    found_month = months_short.get(
                        f'{p1:02d}', found_month
                    )
    if not found_year:
        yy_match = re.search(
            r'[\-_](\d{2})[\-_\.]',
            base_name
        )
        if yy_match:
            yy = int(yy_match.group(1))
            if 20 <= yy <= 30:
                found_year = f"20{yy:02d}"

    if found_month and found_year:
        return f"{found_month} {found_year}"
    elif found_month:
        return found_month

    return None


def period_sheet_name_candidates(filename_period):
    """
    Build likely sheet tab names for a period string.
    e.g. 'February 2026' -> ['Feb 2026', 'February 2026']
    """
    if not filename_period:
        return []
    parts = filename_period.strip().split()
    if len(parts) < 2:
        return [filename_period.strip()]
    year = parts[-1]
    month_part = ' '.join(parts[:-1])
    month_title = month_part.title()
    month_abbr = {
        'January': 'Jan', 'February': 'Feb', 'March': 'Mar',
        'April': 'Apr', 'May': 'May', 'June': 'Jun',
        'July': 'Jul', 'August': 'Aug', 'September': 'Sep',
        'October': 'Oct', 'November': 'Nov', 'December': 'Dec',
    }
    abbr = month_abbr.get(month_title)
    out = [f'{month_title} {year}']
    if abbr:
        out.insert(0, f'{abbr} {year}')
    return out


def match_sheet_to_filename_period(sheet_names, filename_period):
    """
    Pick the sheet whose name matches the report period (full month or
    abbreviated, e.g. February 2026 vs Feb 2026).
    """
    if not filename_period or not sheet_names:
        return None
    fp_lower = filename_period.lower()
    for sn in sheet_names:
        sl = sn.lower().strip()
        if fp_lower in sl or sl in fp_lower:
            return sn
    for cand in period_sheet_name_candidates(filename_period):
        cl = cand.lower()
        for sn in sheet_names:
            if sn.strip().lower() == cl:
                return sn
    return None


SKIP_MATH_VALIDATION = {
    'southbend',
    'south bend',
    'star',
    'gabriel grp',
    'gabriel group',
    'blodgett',
}


def validate_and_score(line_items):
    """
    Validates each line item and adds
    confidence scores and validation flags.
    
    Cross-validates: commission ≈ sale × rate
    """
    validated = []

    for item in line_items:
        item = dict(item)
        issues = []
        confidence = 1.0

        # Check required fields
        if not item.get('po_number'):
            item['po_number'] = None
            issues.append('missing_po')
            confidence -= 0.1

        if not item.get('dealer_name'):
            issues.append('missing_dealer')
            confidence -= 0.15

        if not item.get('invoice_number'):
            issues.append('missing_invoice')
            confidence -= 0.1

        # Validate amounts (clean_amount handles strings like "$10.90")
        sale = clean_amount(item.get('sale_amount'))
        rate = clean_amount(item.get('commission_rate'))
        comm = clean_amount(item.get('commission_amount'))

        # Normalize commission rate
        # Rates like 0.05 mean 5% (decimal fraction)
        # Rates like 0.5 mean 0.5% (valid percentage)
        # Rates like 5.0 mean 5% (already percentage)
        # Threshold: below 0.20 is decimal fraction
        if 0 < rate < 0.20:
            rate = rate * 100
        item['sale_amount'] = sale
        item['commission_rate'] = rate
        item['commission_amount'] = comm

        # Skip math validation for manufacturers with
        # non-standard commission calculations
        mfr = str(
            item.get('manufacturer', '')
        ).lower().strip()
        skip_math = any(
            skip in mfr
            for skip in SKIP_MATH_VALIDATION
        )

        # Cross-validate commission math
        # Skip validation for credits/returns
        # where both sale and commission
        # are negative (legitimate reversal)

        is_credit = comm < 0
        is_reversal = sale < 0 and comm < 0

        if is_reversal:
            # Negative sale with negative commission
            # This is a legitimate reversal
            # Validate the absolute values
            if not skip_math and rate != 0:
                expected = abs(sale) * (rate / 100)
                actual = abs(comm)

                if expected > 0:
                    diff_pct = abs(expected - actual) / expected
                    if diff_pct > 0.75:
                        issues.append(
                            f'commission_mismatch_'
                            f'{round(diff_pct*100, 1)}pct'
                        )
                        confidence -= 0.2

        elif is_credit and sale > 0:
            # Positive sale with negative commission
            # This is unusual - flag it
            issues.append('negative_commission_on_positive_sale')
            confidence -= 0.10

        elif not skip_math and sale != 0 and \
             rate != 0 and comm != 0 and \
             not is_credit:
            # Normal positive commission validation
            expected = abs(sale) * (rate / 100)
            actual = abs(comm)

            if expected > 0:
                diff_pct = abs(expected - actual) / expected
                if diff_pct > 0.75:
                    issues.append(
                        f'commission_mismatch_'
                        f'{round(diff_pct*100, 1)}pct'
                    )
                    confidence -= 0.2
                elif diff_pct > 0.25:
                    issues.append(
                        'commission_calculation_diff'
                    )
                    confidence -= 0.05

        # Validate splits add up to total
        if item.get('has_commission_splits'):
            orig = item.get(
                'commission_origination'
            ) or 0
            spec = item.get(
                'commission_specification'
            ) or 0
            dest = item.get(
                'commission_destination'
            ) or 0
            total = item.get(
                'commission_amount'
            ) or 0

            # Only validate if all three present
            if orig and spec and dest and total:
                split_sum = orig + spec + dest
                diff = abs(split_sum - total)

                if total > 0 and diff / total > 0.02:
                    issues.append(
                        'split_commission_mismatch'
                    )
                    confidence -= 0.15
                    print(f"  Split mismatch: "
                          f"{orig}+{spec}+{dest}"
                          f"={split_sum:.2f} "
                          f"vs total={total:.2f}")

        # Normalize and validate date
        date = item.get('invoice_date')
        normalized_date = normalize_date(date)
        if normalized_date:
            item['invoice_date'] = normalized_date
        elif date:
            issues.append('invalid_date_format')
            confidence -= 0.05
        else:
            issues.append('missing_date')
            confidence -= 0.02
        
        item['confidence'] = round(
            max(0.0, min(1.0, confidence)), 2
        )
        item['validation_issues'] = issues
        item['needs_review'] = (
            confidence < 0.85 or 
            len(issues) > 0
        )
        
        validated.append(item)
    
    return validated


# ─────────────────────────────────────
# PO GROUPING ENGINE
# ─────────────────────────────────────

def normalize_po(po_number):
    """
    Normalize PO number for consistent matching.
    Removes spaces, dashes, leading zeros.
    """
    if not po_number:
        return None
    cleaned = re.sub(r'[\s\-]', '', str(po_number))
    return cleaned.upper()


def group_by_po(validated_items):
    """
    Groups line items by PO number.
    Sums commission and sale amounts per PO.
    Identifies items with no PO for manual handling.
    """
    from collections import defaultdict
    
    groups = defaultdict(lambda: {
        'po_number': None,
        'po_number_normalized': None,
        'dealer_name': None,
        'invoice_numbers': [],
        'invoice_dates': [],
        'total_sale_amount': 0.0,
        'total_commission': 0.0,
        'line_item_count': 0,
        'line_items': [],
        'has_review_items': False,
        'all_issues': [],
        'min_confidence': 1.0,
        'commission_origination': None,
        'commission_specification': None,
        'commission_destination': None,
        'has_commission_splits': False,
        'commission_rebate': None,
    })
    
    no_po_items = []
    
    for item in validated_items:
        po = item.get('po_number')
        
        if not po:
            no_po_items.append(item)
            continue
        
        po_norm = normalize_po(po)
        g = groups[po_norm]
        
        g['po_number'] = po
        g['po_number_normalized'] = po_norm
        g['dealer_name'] = (
            g['dealer_name'] or 
            item.get('dealer_name')
        )
        g['total_sale_amount'] += (
            item.get('sale_amount') or 0
        )
        g['total_commission'] += (
            item.get('commission_amount') or 0
        )
        g['line_item_count'] += 1
        g['line_items'].append(item)

        # Sum split commissions (Orig, Spec, Dest) when multiple line items
        # share the same PO - e.g. Blodgett sends one row per line item
        orig = item.get('commission_origination') or 0
        spec = item.get('commission_specification') or 0
        dest = item.get('commission_destination') or 0
        if orig or spec or dest:
            g['commission_origination'] = (
                (g.get('commission_origination') or 0) + orig
            )
            g['commission_specification'] = (
                (g.get('commission_specification') or 0) + spec
            )
            g['commission_destination'] = (
                (g.get('commission_destination') or 0) + dest
            )
            g['has_commission_splits'] = True

        if item.get('has_commission_splits'):
            g['has_commission_splits'] = True

        rebate_val = item.get('commission_rebate') or 0
        if rebate_val:
            g['commission_rebate'] = (
                g.get('commission_rebate') or 0
            ) + rebate_val

        inv = item.get('invoice_number')
        if inv and inv not in g['invoice_numbers']:
            g['invoice_numbers'].append(inv)
        
        date = item.get('invoice_date')
        if date and date not in g['invoice_dates']:
            g['invoice_dates'].append(date)
        
        if item.get('needs_review'):
            g['has_review_items'] = True
        
        g['all_issues'].extend(
            item.get('validation_issues', [])
        )
        
        conf = item.get('confidence', 1.0)
        if conf < g['min_confidence']:
            g['min_confidence'] = conf
    
    # Round totals and split values
    for po_norm, g in groups.items():
        g['total_sale_amount'] = round(
            g['total_sale_amount'], 2
        )
        g['total_commission'] = round(
            g['total_commission'] or 0, 2
        )
        g['min_confidence'] = round(
            g['min_confidence'], 2
        )
        for sf in [
            'commission_origination',
            'commission_specification',
            'commission_destination',
            'commission_rebate'
        ]:
            if g.get(sf) is not None:
                g[sf] = round(g[sf], 2)
        # Ensure has_commission_splits when any split is non-zero
        if (g.get('commission_origination') or 0) or (
            g.get('commission_specification') or 0
        ) or (g.get('commission_destination') or 0):
            g['has_commission_splits'] = True
        g['needs_review'] = (
            g['has_review_items'] or
            g['min_confidence'] < 0.85
        )

    def _norm_inv(v):
        s = str(v).strip() if v is not None else ''
        if not s:
            return ''
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
            return s
        except (ValueError, TypeError):
            return s

    # Collect all invoice numbers already in PO groups
    po_invoice_numbers = set()
    for group in groups.values():
        for inv in group.get('invoice_numbers', []):
            n = _norm_inv(inv)
            if n:
                po_invoice_numbers.add(n)

    # Star reports repeat the same invoice on multiple rows (Orig/Spec/Dest);
    # do not treat same-invoice no-PO rows as duplicates of PO rows.
    star_mfr = any(
        str(it.get('manufacturer') or '').lower().strip() == 'star'
        for it in validated_items
    )

    # Filter no-PO items - exclude any that share an invoice number
    # with a PO group as these are likely duplicates
    filtered_no_po = []
    duplicate_no_po = []

    if star_mfr:
        filtered_no_po = list(no_po_items)
    else:
        for item in no_po_items:
            inv_norm = _norm_inv(item.get('invoice_number'))
            if inv_norm and inv_norm in po_invoice_numbers:
                duplicate_no_po.append(item)
            else:
                filtered_no_po.append(item)

        if duplicate_no_po:
            print(f"  Excluded {len(duplicate_no_po)} "
                  f"no-PO items as duplicates of "
                  f"existing PO groups")

    return {
        'grouped': dict(groups),
        'no_po_items': filtered_no_po,
        'total_pos': len(groups),
        'total_no_po': len(filtered_no_po),
        'total_items': len(validated_items),
        'items_needing_review': sum(
            1 for item in validated_items
            if item.get('needs_review')
        )
    }


# ─────────────────────────────────────
# TEMPLATE-BASED EXTRACTION (Excel/CSV)
# ─────────────────────────────────────

def _parse_sheet_with_header_detection(
    xl, sheet_name, mapped_col_names
):
    """
    Parse Excel sheet, detecting header row if default (0) fails.
    Some reports have metadata in row 0; headers in a later row.
    """
    import pandas as pd

    def _cell_matches(col_name, cells):
        if not col_name:
            return False
        cn = str(col_name).strip().lower()
        for c in cells:
            if c is None:
                continue
            try:
                if hasattr(c, '__float__') and pd.isna(c):
                    continue
            except (TypeError, ValueError):
                pass
            val = str(c).strip().lower()
            if cn in val or val in cn:
                return True
        return False

    # Try default header=0 first
    df = xl.parse(sheet_name)
    available = set(str(c) for c in df.columns)
    for mn in mapped_col_names:
        if not mn:
            continue
        mn_str = str(mn).strip()
        if mn_str in available:
            return df
        mn_lower = mn_str.lower()
        for c in available:
            if mn_lower in str(c).lower():
                return df

    # No match: try finding header row (first 25 rows)
    df_raw = xl.parse(sheet_name, header=None)
    header_row = None

    def _cell_matches_short(col_name, cells, max_len=50):
        """Match only if a cell is short (actual header, not report title)."""
        if not col_name:
            return False
        cn = str(col_name).strip().lower()
        for c in cells:
            if c is None:
                continue
            try:
                if hasattr(c, '__float__') and pd.isna(c):
                    continue
            except (TypeError, ValueError):
                pass
            val = str(c).strip()
            if len(val) > max_len:
                continue
            val_lower = val.lower()
            if cn in val_lower or val_lower in cn:
                return True
        return False

    for idx in range(min(25, len(df_raw))):
        row = df_raw.iloc[idx]
        cells = [row.iloc[i] for i in range(len(row))]
        for mn in mapped_col_names:
            if _cell_matches(mn, cells):
                header_row = idx
                break
        if header_row is None:
            for kw in ['comm $', 'commission base', 'commission amount']:
                if _cell_matches_short(kw, cells):
                    header_row = idx
                    break
        if header_row is not None:
            break

    if header_row is not None:
        df = xl.parse(sheet_name, header=header_row)
        print(f"  Detected header row {header_row} in '{sheet_name}'")
        return df
    return df_raw


def _parse_dataframe_header_detection(df, mapped_col_names):
    """
    Same header-row logic as _parse_sheet_with_header_detection,
    but for an in-memory DataFrame (e.g. from SLK parse).
    """
    import pandas as pd

    if df is None or df.empty:
        return df

    def _cell_matches(col_name, cells):
        if not col_name:
            return False
        cn = str(col_name).strip().lower()
        for c in cells:
            if c is None:
                continue
            try:
                if hasattr(c, '__float__') and pd.isna(c):
                    continue
            except (TypeError, ValueError):
                pass
            val = str(c).strip().lower()
            if cn in val or val in cn:
                return True
        return False

    def _cell_matches_short(col_name, cells, max_len=50):
        if not col_name:
            return False
        cn = str(col_name).strip().lower()
        for c in cells:
            if c is None:
                continue
            try:
                if hasattr(c, '__float__') and pd.isna(c):
                    continue
            except (TypeError, ValueError):
                pass
            val = str(c).strip()
            if len(val) > max_len:
                continue
            val_lower = val.lower()
            if cn in val_lower or val_lower in cn:
                return True
        return False

    mapped_col_names = [m for m in (mapped_col_names or []) if m]
    available = set(str(c) for c in df.columns)
    for mn in mapped_col_names:
        if not mn:
            continue
        mn_str = str(mn).strip()
        if mn_str in available:
            return df
        mn_lower = mn_str.lower()
        for c in available:
            if mn_lower in str(c).lower():
                return df

    df_raw = df.copy()
    df_raw.columns = range(df_raw.shape[1])
    header_row = None
    for idx in range(min(25, len(df_raw))):
        row = df_raw.iloc[idx]
        cells = [row.iloc[i] for i in range(len(row))]
        for mn in mapped_col_names:
            if _cell_matches(mn, cells):
                header_row = idx
                break
        if header_row is None:
            for kw in ['comm $', 'commission base', 'commission amount']:
                if _cell_matches_short(kw, cells):
                    header_row = idx
                    break
        if header_row is not None:
            break

    if header_row is not None:
        hdr = df_raw.iloc[header_row].astype(str).str.strip()
        hdr = [h if h and h.lower() != 'nan' else f'Column{i}'
               for i, h in enumerate(hdr)]
        body = df_raw.iloc[header_row + 1:].copy()
        body.columns = hdr
        print(f"  Detected header row {header_row} (SLK/DataFrame)")
        return body
    return df


def _load_slk_sheets_for_template(file_path, mapping):
    """
    Load SYLK (.slk) as sheet dicts compatible with
    extract_excel_with_template (single sheet typical).
    """
    import pandas as pd

    try:
        xl = pd.ExcelFile(file_path, engine='xlrd')
        primary = (mapping.get('primary_sheet') or '').strip()
        out = []
        for sn in xl.sheet_names:
            if primary and sn.lower() != primary.lower():
                continue
            df = xl.parse(sn)
            if df is not None and not df.empty:
                out.append({'name': sn, 'df': df})
        if out:
            return out
    except Exception as e:
        print(f"  SLK via xlrd/ExcelFile: {e}")

    df = load_slk_dataframe(file_path)
    if df is None or df.empty:
        return []
    sheet_name = (
        mapping.get('primary_sheet') or 'Sheet1'
    ).strip() or 'Sheet1'
    return [{'name': sheet_name, 'df': df}]


def extract_excel_with_template(
    file_path, file_name, mapping
):
    """
    Extract commission records from Excel
    using a column mapping template.
    Uses pandas directly - zero AI calls.
    """
    import pandas as pd
    from pathlib import Path

    manufacturer = mapping.get('manufacturer', '')
    period = mapping.get('period', '')
    skip_pattern = (
        mapping.get('skip_rows_where', '') or ''
    ).lower()

    # Build skip keywords from pattern
    skip_keywords = []
    for word in ['total', 'subtotal', 'grand',
                 'summary', 'page', 'report']:
        if word in skip_pattern or True:
            skip_keywords.append(word)

    ext = Path(file_path).suffix.lower()
    all_records = []

    if ext in ['.csv', '.tsv']:
        sep = '\t' if ext == '.tsv' else ','
        try:
            df = pd.read_csv(file_path, sep=sep)
        except Exception as e:
            print(f"  CSV read error: {e}")
            return []
        if df.empty:
            return []
        sheets_data = [{'name': 'Sheet1', 'df': df}]
    elif ext == '.slk':
        raw_sheets = _load_slk_sheets_for_template(
            file_path, mapping
        )
        if not raw_sheets:
            print("  SLK: could not load any sheet")
            return []
        mapped_cols = [
            mapping.get('commission_amount'),
            mapping.get('po_number'),
            mapping.get('dealer_name'),
            mapping.get('comm_credit'),
            mapping.get('invoice_date'),
            mapping.get('po_date'),
        ]
        sheets_data = []
        for item in raw_sheets:
            df_h = _parse_dataframe_header_detection(
                item['df'], mapped_cols
            )
            if df_h is not None and not df_h.empty:
                sheets_data.append({
                    'name': item['name'],
                    'df': df_h,
                })
        if not sheets_data:
            print("  SLK: no data after header detection")
            return []
    else:
        xl = pd.ExcelFile(file_path)

        # Determine which sheets to process
        primary_sheet = mapping.get('primary_sheet')
        skip_sheets = mapping.get('skip_sheets', []) or []

        # If filename has period (e.g. February 2026), prefer sheet matching it
        # (Star uses tabs like 'Feb 2026' while the filename says 'February 2026')
        filename_period = extract_period_from_filename(file_name)
        if filename_period:
            resolved = match_sheet_to_filename_period(
                xl.sheet_names, filename_period
            )
            if resolved:
                primary_sheet = resolved
                print(
                    f"  Using sheet '{resolved}' "
                    f"(matches filename period)"
                )

        # Normalize for comparison
        skip_lower = [s.lower().strip() for s in skip_sheets if s]

        sheets_to_process = []
        for sheet_name in xl.sheet_names:
            if sheet_name.lower() in skip_lower:
                print(f"  Skipping sheet: {sheet_name} "
                      f"(in skip_sheets)")
                continue
            if primary_sheet and sheet_name.lower() != primary_sheet.lower().strip():
                print(f"  Skipping sheet: {sheet_name} "
                      f"(not primary sheet)")
                continue
            sheets_to_process.append(sheet_name)

        if not sheets_to_process:
            print("  No primary sheet found - "
                  "processing all sheets")
            sheets_to_process = xl.sheet_names

        print(f"  Processing sheets: {sheets_to_process}")

        sheets_data = []
        mapped_cols = [
            mapping.get('commission_amount'),
            mapping.get('po_number'),
            mapping.get('dealer_name'),
            mapping.get('comm_credit'),
            mapping.get('invoice_date'),
            mapping.get('po_date'),
        ]
        for sn in sheets_to_process:
            try:
                df = _parse_sheet_with_header_detection(
                    xl, sn, mapped_cols
                )
                if df is not None and not df.empty:
                    sheets_data.append({'name': sn, 'df': df})
            except Exception as e:
                print(f"  Sheet {sn} error: {e}")

    for item in sheets_data:
        sheet_name = item['name']
        df = item['df']
        try:
            if df.empty:
                continue
        except Exception:
            continue

        print(f"  Sheet '{sheet_name}': "
              f"{len(df)} rows, "
              f"{len(df.columns)} columns")
        print(f"  Available columns: "
              f"{list(df.columns)[:10]}")
        print(f"  Looking for comm_col: "
              f"'{mapping.get('commission_amount')}'")

        # Map column names
        po_col = mapping.get('po_number')
        dealer_col = mapping.get('dealer_name')
        inv_col = mapping.get('invoice_number')
        date_col = mapping.get('invoice_date')
        sale_col = mapping.get('sale_amount')
        rate_col = mapping.get('commission_rate')
        comm_col = mapping.get('commission_amount')

        # Find columns that exist
        available = set(df.columns.astype(str))

        def _norm(s):
            return re.sub(r'\s+', ' ', str(s or '').strip().lower())

        def safe_get(col):
            if not col:
                return None
            col_str = str(col).strip()
            if col_str in available:
                return col_str
            col_norm = _norm(col)
            for c in available:
                if _norm(c) == col_norm:
                    return str(c)
            return None

        po_col = safe_get(po_col)
        dealer_col = safe_get(dealer_col)
        inv_col = safe_get(inv_col)
        date_col = safe_get(date_col)
        po_date_col = safe_get(mapping.get('po_date'))
        sale_col = safe_get(sale_col)
        rate_col = safe_get(rate_col)
        comm_col = safe_get(comm_col)

        # Get split commission column mappings
        orig_col = safe_get(
            mapping.get('commission_origination')
        )
        spec_col = safe_get(
            mapping.get('commission_specification')
        )
        dest_col = safe_get(
            mapping.get('commission_destination')
        )
        rebate_col = safe_get(
            mapping.get('commission_rebate')
        )
        comm_credit_col = safe_get(
            mapping.get('comm_credit')
        )
        mfr_lower = str(manufacturer or '').lower().strip()

        if not comm_col:
            # Better fallback: find columns with numeric dollar-like values
            for col in df.columns:
                col_str = str(col).lower()
                # Skip obvious non-commission cols
                if any(skip in col_str for skip in [
                    'sale', 'amount', 'price',
                    'credit', 'total', 'report',
                    'date', 'name', 'po', 'invoice',
                    'rate', 'percent', '%'
                ]) and 'commission' not in col_str and 'comm' not in col_str:
                    continue
                # Check if col has numeric values
                try:
                    numeric_count = pd.to_numeric(
                        df[col], errors='coerce'
                    ).notna().sum()
                    if numeric_count > len(df) * 0.3:
                        vals = pd.to_numeric(
                            df[col], errors='coerce'
                        ).dropna()
                        if len(vals) > 0:
                            median = vals.median()
                            if 0.01 < abs(median) < 100000:
                                comm_col = str(col)
                                print(f"  Auto-detected "
                                      f"commission col: "
                                      f"'{comm_col}'")
                                break
                except Exception:
                    continue
        if not comm_col:
            print(f"  No commission column found in sheet {sheet_name}")
            continue

        def _is_missing_date_val(v):
            if v is None:
                return True
            try:
                if pd.isna(v):
                    return True
            except (TypeError, ValueError):
                pass
            if isinstance(v, str) and not v.strip():
                return True
            return False

        for _, row in df.iterrows():
            comm_val = row.get(comm_col)
            if comm_val is None or pd.isna(comm_val):
                continue

            comm = clean_amount(comm_val)
            star_credit_rows = (
                mfr_lower == 'star' and comm_credit_col
            )
            if comm == 0 and not star_credit_rows:
                continue

            # Skip total/summary rows
            row_str = ' '.join([
                str(v).lower()
                for v in row.values
                if v is not None
                and not pd.isna(v)
            ])

            label_val = ''
            if po_col and row.get(po_col):
                label_val = str(
                    row.get(po_col)
                ).lower()
            elif dealer_col and row.get(dealer_col):
                label_val = str(
                    row.get(dealer_col)
                ).lower()

            if any(
                kw in label_val
                for kw in ['total', 'subtotal',
                           'grand total', 'summary']
            ):
                continue

            def get_val(col):
                if not col:
                    return None
                v = row.get(col)
                if v is None:
                    return None
                try:
                    if pd.isna(v):
                        return None
                except (TypeError, ValueError):
                    pass
                return v

            raw_date = get_val(date_col)
            if _is_missing_date_val(raw_date) and po_date_col:
                raw_date = get_val(po_date_col)

            invoice_date = None
            if not _is_missing_date_val(raw_date):
                if hasattr(raw_date, 'strftime'):
                    invoice_date = raw_date.strftime(
                        '%Y-%m-%d'
                    )
                else:
                    invoice_date = normalize_date(
                        raw_date
                    )

            po_str = str(get_val(po_col) or '').strip()
            inv_str = str(get_val(inv_col) or '').strip()
            po_source_auto = None
            # If no PO but ORDER # / invoice number is present (e.g. Cambro
            # distributor rows), use ORDER # as PO so rows are included in
            # totals and grouping. Do not require manufacturer === Cambro
            # (mapping may omit manufacturer on some uploads).
            if not po_str and inv_str:
                po_str = inv_str
                po_source_auto = 'invoice_used_as_po'

            record = {
                'po_number': po_str or None,
                'dealer_name': (
                    str(get_val(dealer_col)).strip()
                    if get_val(dealer_col) else None
                ),
                'invoice_number': inv_str or None,
                'invoice_date': invoice_date,
                'sale_amount': clean_amount(
                    get_val(sale_col)
                ),
                'commission_rate': clean_amount(
                    get_val(rate_col)
                ),
                'commission_amount': comm,
                'manufacturer': manufacturer,
                'period': period
            }
            if po_source_auto:
                record['po_source'] = po_source_auto

            # Skip grand total / summary rows
            dealer_val = str(record.get('dealer_name') or '').strip()
            po_val = str(record.get('po_number') or '').strip()
            inv_val = str(record.get('invoice_number') or '').strip()

            # Skip if commission exists but dealer, PO, and invoice
            # are ALL empty (grand total row)
            if (comm > 0 and
                    not dealer_val and
                    not po_val and
                    not inv_val):
                print(f'  Skipping grand total row: comm={comm}')
                continue

            # Extract split values: dedicated Orig/Spec/Dest columns,
            # or Star-style "Comm Credit" row type (one amount per row).
            origination = None
            specification = None
            destination = None
            used_comm_credit = False

            if comm_credit_col:
                cc_raw = get_val(comm_credit_col)
                cc = str(cc_raw or '').strip()
                record['comm_credit_type'] = cc or None
                if cc:
                    used_comm_credit = True
                    ccl = cc.lower()
                    if 'origination' in ccl:
                        origination = comm
                        specification = 0
                        destination = 0
                    elif 'specification' in ccl:
                        origination = 0
                        specification = comm
                        destination = 0
                    elif 'destination' in ccl:
                        origination = 0
                        specification = 0
                        destination = comm
            else:
                record['comm_credit_type'] = None

            if not used_comm_credit:
                if orig_col:
                    v = get_val(orig_col)
                    if v is not None:
                        origination = clean_amount(v) or None

                if spec_col:
                    v = get_val(spec_col)
                    if v is not None:
                        specification = clean_amount(v) or None

                if dest_col:
                    v = get_val(dest_col)
                    if v is not None:
                        destination = clean_amount(v) or None

            # Has splits if at least one is non-zero
            has_splits = any([
                origination,
                specification,
                destination
            ])

            record['commission_origination'] = origination
            record['commission_specification'] = specification
            record['commission_destination'] = destination
            record['has_commission_splits'] = has_splits

            rebate = None
            if rebate_col:
                v = get_val(rebate_col)
                if v is not None:
                    rebate = clean_amount(v) or None

            record['commission_rebate'] = rebate

            # Star: same Net Sales is repeated on Orig/Spec/Dest rows; only
            # the Origination row should contribute to sale totals per line.
            if (
                mfr_lower == 'star'
                and used_comm_credit
                and record.get('comm_credit_type')
            ):
                cct = record['comm_credit_type'].lower()
                if (
                    'specification' in cct
                    or 'destination' in cct
                ):
                    record['sale_amount'] = 0.0

            # Skip rows that are voucher/accounting entries or summary
            dealer = str(
                record.get('dealer_name') or ''
            ).upper()
            skip_dealers = [
                'VOUCHER', 'REBATE', 'ADJUSTMENT',
                'TOTAL', 'SUBTOTAL', 'GRAND TOTAL',
                'SUMMARY', 'TRANSFER', 'TOTAL COMMISSION'
            ]
            if any(skip in dealer for skip in skip_dealers):
                print(f'  Skipping summary row: dealer={dealer_val}')
                continue

            all_records.append(record)

    print(f"  Extracted {len(all_records)} records "
          f"(0 AI calls for data)")
    return all_records


# ─────────────────────────────────────
# MAIN EXTRACTION PIPELINE
# ─────────────────────────────────────

def process_commission_file(
    file_path,
    db_path=None,
    check_duplicates=True,
    original_filename=None
):
    from pathlib import Path

    path = Path(file_path)
    ext = path.suffix.lower()
    file_name = path.name
    # Use original filename for period extraction when file
    # was saved with temp name (e.g. API upload)
    period_filename = original_filename or file_name

    print(f"\nProcessing: {file_name}")
    print("-" * 50)

    try:
        # Step 1: Read file for hash/dedup
        print("Step 1: Reading file...")
        doc = read_commission_file(file_path)
        print(f"  {doc['char_count']} chars "
              f"({doc['file_type']})")

        # Step 2: Check duplicates
        if check_duplicates and db_path:
            if is_duplicate(doc['hash'], db_path):
                print("  DUPLICATE - skipping")
                return {
                    'document': doc,
                    'duplicate': True,
                    'status': 'duplicate',
                    'error': None
                }

        # Step 3: Extract based on file type
        if ext in ['.xlsx', '.xlsm', '.xls', '.xlsb', '.slk']:

            print("Step 2: Excel extraction "
                  "(template-based)...")

            from parser.commission_mapper import (
                get_or_create_template,
                extract_manufacturer_from_filename
            )

            manufacturer = (
                extract_manufacturer_from_filename(
                    period_filename
                )
            )
            print(f"  Manufacturer: {manufacturer}")

            if not db_path:
                db_path = os.path.join(
                    os.path.dirname(__file__),
                    '..', 'data', 'platform.db'
                )

            mapping = get_or_create_template(
                file_path, file_name,
                manufacturer, db_path,
                original_filename=period_filename,
            )

            filename_period = extract_period_from_filename(
                period_filename
            )
            template_period = mapping.get('period')
            print(f"  Filename period: {filename_period}")
            print(f"  Template period: {template_period}")
            print(f"  Using: {filename_period or template_period}")
            if filename_period:
                mapping['period'] = filename_period

            line_items = extract_excel_with_template(
                file_path, period_filename, mapping
            )

            extraction = {
                'manufacturer': mapping.get(
                    'manufacturer'
                ) or manufacturer,
                'period': mapping.get('period', ''),
                'currency': 'USD',
                'line_items': line_items,
                'total_extracted': len(line_items),
                '_source': 'excel'
            }

        elif ext in ['.csv', '.tsv']:

            print("Step 2: CSV extraction "
                  "(template-based)...")

            from parser.commission_mapper import (
                get_or_create_template,
                extract_manufacturer_from_filename
            )

            manufacturer = (
                extract_manufacturer_from_filename(
                    period_filename
                )
            )

            if not db_path:
                db_path = os.path.join(
                    os.path.dirname(__file__),
                    '..', 'data', 'platform.db'
                )

            mapping = get_or_create_template(
                file_path, file_name,
                manufacturer, db_path,
                original_filename=period_filename,
            )

            filename_period = extract_period_from_filename(
                period_filename
            )
            template_period = mapping.get('period')
            print(f"  Filename period: {filename_period}")
            print(f"  Template period: {template_period}")
            print(f"  Using: {filename_period or template_period}")
            if filename_period:
                mapping['period'] = filename_period

            line_items = extract_excel_with_template(
                file_path, period_filename, mapping
            )

            extraction = {
                'manufacturer': mapping.get(
                    'manufacturer'
                ) or manufacturer,
                'period': mapping.get('period', ''),
                'currency': 'USD',
                'line_items': line_items,
                'total_extracted': len(line_items),
                '_source': 'csv'
            }

        else:
            # PDF and TXT: use LLM extraction
            print("Step 2: LLM extraction "
                  "(PDF/TXT)...")
            extraction = extract_with_gemini(
                doc['content'], file_name
            )

            filename_period = extract_period_from_filename(
                period_filename
            )
            print(f"  Filename period: {filename_period}")
            print(f"  Template period: {extraction.get('period')}")
            print(f"  Using: {filename_period or extraction.get('period')}")
            if filename_period:
                extraction['period'] = filename_period
            extraction['_source'] = 'pdf'

        print(f"  Extracted: "
              f"{extraction['total_extracted']} "
              f"line items")

        # Step 4: Validate and score
        print("Step 3: Validating...")
        validated = validate_and_score(
            extraction['line_items']
        )

        needs_review = sum(
            1 for i in validated
            if i.get('needs_review')
        )
        print(f"  {len(validated)} items, "
              f"{needs_review} need review")

        # Step 5: Group by PO
        print("Step 4: Grouping by PO...")
        grouped = group_by_po(validated)
        print(f"  {grouped['total_pos']} POs, "
              f"{grouped['total_no_po']} without PO")

        # Step 6: Summary (include no-PO items in total)
        po_total = sum(
            g['total_commission']
            for g in grouped['grouped'].values()
        )
        no_po_total = sum(
            item.get('commission_amount') or 0
            for item in grouped['no_po_items']
        )
        # Exclude no-PO items that are likely summary/total rows
        # (e.g. Excel Grand Total with no PO but full commission)
        if no_po_total > 0 and po_total > 0:
            if abs(no_po_total - po_total) / po_total < 0.02:
                no_po_total = 0  # no-PO sum ≈ PO sum → summary row
            else:
                # Exclude individual items that look like totals
                for item in grouped['no_po_items']:
                    amt = item.get('commission_amount') or 0
                    if amt >= 0.9 * po_total:
                        no_po_total -= amt
            # For Excel/CSV: no-PO items often indicate duplicate rows
            # or data quality issues. Only add if minimal (< 5% of PO total).
            src = extraction.get('_source', '')
            if src in ('excel', 'csv') and no_po_total > 0.05 * po_total:
                no_po_total = 0
        total_commission = po_total + no_po_total

        summary = {
            'total_pos': grouped['total_pos'],
            'total_no_po': grouped['total_no_po'],
            'total_line_items': (
                grouped['total_items']
            ),
            'items_needing_review': (
                grouped['items_needing_review']
            ),
            'total_commission': round(
                total_commission, 2
            ),
            'manufacturer': extraction['manufacturer'],
            'period': extraction.get('period', '')
        }

        print(f"\nSummary:")
        print(f"  Commission: "
              f"${summary['total_commission']:,.2f}")
        print(f"  POs: {summary['total_pos']}")
        print(f"  Needs review: "
              f"{summary['items_needing_review']}")

        return {
            'document': doc,
            'extraction': extraction,
            'grouped': grouped['grouped'],
            'no_po_items': grouped['no_po_items'],
            'summary': summary,
            'duplicate': False,
            'status': 'success',
            'error': None
        }

    except Exception as e:
        import traceback
        print(f"ERROR: {e}")
        traceback.print_exc()
        return {
            'document': {
                'file_name': file_name
            },
            'duplicate': False,
            'status': 'error',
            'error': str(e)
        }


def is_duplicate(doc_hash, db_path):
    """Check if document was already processed."""
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS 
            processed_documents (
                hash TEXT PRIMARY KEY,
                file_name TEXT,
                processed_at TEXT
            )
        """)
        
        cursor.execute(
            "SELECT hash FROM processed_documents "
            "WHERE hash = ?",
            (doc_hash,)
        )
        
        exists = cursor.fetchone() is not None
        conn.close()
        return exists
    
    except Exception:
        return False


def mark_as_processed(doc_hash, file_name, db_path):
    """Mark document as processed to prevent reprocessing."""
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS 
            processed_documents (
                hash TEXT PRIMARY KEY,
                file_name TEXT,
                processed_at TEXT
            )
        """)
        
        cursor.execute("""
            INSERT OR IGNORE INTO processed_documents
            (hash, file_name, processed_at)
            VALUES (?, ?, ?)
        """, (
            doc_hash, 
            file_name,
            datetime.now().isoformat()
        ))
        
        conn.commit()
        conn.close()
    
    except Exception as e:
        print(f"Warning: could not mark as processed: {e}")
